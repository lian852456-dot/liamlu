#!/usr/bin/env python3
import html
import hashlib
import json
import math
import os
import re
import shutil
import zipfile
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from pathlib import Path
from xml.etree import ElementTree as ET

BASE = Path("/Users/liamlu/Downloads")
PROJECT_ROOT = Path("/Users/liamlu/Downloads/liam-agent")
OUT_DIR = Path(os.environ.get(
    "PHONE_AWARDS_OUTPUT_DIR",
    "/Users/liamlu/Downloads/liam-agent/report-automation/outputs",
)).expanduser()
CONFIG_DIR = PROJECT_ROOT / "report-automation/config"
PHONE_INPUT_DIR = PROJECT_ROOT / "report-automation/input/google-drive/phone-awards"
PHONE_AWARDS_ORIGIN_DIR = Path(os.environ.get(
    "PHONE_AWARDS_ORIGIN_DIR",
    "/Users/liamlu/Library/CloudStorage/OneDrive-個人/TWM每日戰報",
)).expanduser()
PHONE_AWARDS_LOG_DIR = Path(os.environ.get(
    "PHONE_AWARDS_LOG_DIR",
    str(PROJECT_ROOT / "report-automation/logs"),
)).expanduser()
Y26_FALLBACK_SRC = BASE / "Y2606+%E9%87%8D%E9%BB%9E"
STORE_FALLBACK_SRC = BASE / "01-08-03-%28%E5%AF%86%29%E7%9B%B4%E7%87%9F_%E6%89%8B%E6%A9%9F%E7%AB%B6%E8%B3%BD%E6%97%A5%E5%A0%B1_%E5%BA%97%E9%BB%9E%E9%81%94%E6%88%90%E7%8E%87%E3%80%81%E6%8E%92%E5%90%8D%E5%8F%8A%E7%8D%8E%E9%87%91"
PERSON_FALLBACK_SRC = BASE / "01-08-04-%28%E5%AF%86%29%E7%9B%B4%E7%87%9F_%E6%89%8B%E6%A9%9F%E7%AB%B6%E8%B3%BD%E6%97%A5%E5%A0%B1_%E5%80%8B%E4%BA%BA%E9%81%94%E6%88%90%E7%8E%87%E3%80%81%E6%8E%92%E5%90%8D%E5%8F%8A%E7%8D%8E%E9%87%91"
Y26_SRC = Y26_FALLBACK_SRC
STORE_SRC = STORE_FALLBACK_SRC
PERSON_SRC = PERSON_FALLBACK_SRC
EXACT_STORE_SOURCE = os.environ.get("PHONE_AWARDS_STORE_SOURCE", "").strip()
EXACT_PERSON_SOURCE = os.environ.get("PHONE_AWARDS_PERSON_SOURCE", "").strip()
PHONE_AWARDS_SOURCE_MODE = os.environ.get("PHONE_AWARDS_SOURCE_MODE", "onedrive-cloud").strip()
PHONE_AWARDS_CLOUD_PROVIDERS = {"onedrive-cloud", "google-drive-cloud"}
PHONE_AWARDS_LOCAL_EMERGENCY_ENABLED = os.environ.get("PHONE_AWARDS_LOCAL_EMERGENCY_ENABLED", "") == "1"
PHONE_AWARDS_CLOUD_MANIFEST_PATH = os.environ.get("PHONE_AWARDS_CLOUD_MANIFEST", "").strip()
REPORT_RUN_DATE = os.environ.get("REPORT_DATE_ISO", "").strip()
TODAY_REPORT_DATA_PATH = Path(os.environ.get(
    "REPORT_DATA_OUTPUT",
    str(PROJECT_ROOT / "report-automation/work/today_report_data.json"),
)).expanduser()
VISIBLE_STORE_ITEMS = [
    "Pixel10/10 Pro/10 Pro XL/10a",
    "razrfold",
    "S26Ultra/ZFold8/ZFold8Ultra",
    "R11",
    "X300/V70FE",
    "Pixel11Pro/11ProXL/11ProFold",
    "S26/S26+/ZFlip8",
    "Pixel11",
    "Reno16F",
    "A57",
    "A6x",
    "A27/A17",
    "Y21",
]
ITEM_ALIASES = {
    "Pixel10/10Pro/10ProXL/10a": "Pixel10/10 Pro/10 Pro XL/10a",
    "X300/X300Pro/V70FE": "X300/V70FE",
    "1VIII/razrfold": "razrfold",
}
MANAGER_AWARD_RULES = {
    "Pixel10/10 Pro/10 Pro XL/10a": [770, 985, 1150, 1315, 1645],
    "razrfold": [240, 310, 755, 860, 1080],
    "S26Ultra/ZFold8/ZFold8Ultra": [560, 720, 845, 965, 1205],
    "R11": [385, 495, 580, 665, 830],
    "X300/V70FE": [2130, 2740, 3195, 3655, 4565],
    "Pixel11Pro/11ProXL/11ProFold": [255, 325, 380, 435, 545],
    "S26/S26+/ZFlip8": [570, 735, 855, 980, 1220],
    "Pixel11": [235, 305, 355, 405, 505],
    "Reno16F": [1065, 1375, 1595, 1825, 2280],
    "A57": [1615, 2075, 2420, 2770, 3460],
    "A6x": [640, 820, 955, 1095, 1370],
    "A27/A17": [285, 370, 430, 495, 615],
    "Y21": [20, 25, 30, 35, 45],
}
SUPERVISOR_AWARD_RULES = {
    "Pixel10/10 Pro/10 Pro XL/10a": [800, 980, 1230, 1475, 1840],
    "razrfold": [250, 305, 840, 1010, 1260],
    "S26Ultra/ZFold8/ZFold8Ultra": [585, 720, 900, 1080, 1350],
    "R11": [400, 495, 620, 745, 930],
    "X300/V70FE": [2215, 2725, 3410, 4090, 5115],
    "Pixel11Pro/11ProXL/11ProFold": [265, 325, 410, 490, 615],
    "S26/S26+/ZFlip8": [595, 735, 920, 1105, 1380],
    "Pixel11": [245, 300, 375, 455, 565],
    "Reno16F": [1115, 1375, 1720, 2060, 2575],
    "A57": [1695, 2085, 2610, 3130, 3910],
    "A6x": [655, 805, 1010, 1210, 1510],
    "A27/A17": [295, 365, 455, 545, 680],
    "Y21": [20, 25, 35, 40, 50],
}
AWARD_RULE_BLOCKS = [
    {"header_row": 1, "manager_row": 3, "supervisor_row": 5, "starts": ["F", "J", "N", "R", "V"]},
    {"header_row": 17, "manager_row": 19, "supervisor_row": 21, "starts": ["F", "J", "N", "R", "V"]},
]
MODEL_DISPLAY_NAMES = {}
ACTIVE_AWARD_CONFIG = None
CLOUD_SOURCE_MANIFEST = None

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"

ET.register_namespace("", NS)
ET.register_namespace("r", REL_NS)
ET.register_namespace("", PKG_REL_NS)
ET.register_namespace("", CONTENT_NS)


def q(tag, ns=NS):
    return f"{{{ns}}}{tag}"


def col_to_num(col):
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n


def num_to_col(n):
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def cell_ref(col, row):
    return f"{num_to_col(col)}{row}"


def max_cell_col(cells, fallback):
    max_col = fallback
    for ref in cells:
        match = re.match(r"([A-Z]+)\d+", ref or "")
        if match:
            max_col = max(max_col, col_to_num(match.group(1)))
    return max_col


def read_shared_strings(base):
    path = base / "xl/sharedStrings.xml"
    if not path.exists():
        return []
    root = ET.parse(path).getroot()
    return ["".join(t.text or "" for t in si.iter(q("t"))) for si in root.findall(q("si"))]


def text_of_cell(cell, shared_strings):
    if cell is None:
        return ""
    v = cell.find(q("v"))
    if v is None:
        inline = cell.find(q("is"))
        if inline is not None:
            return "".join(t.text or "" for t in inline.iter(q("t")))
        return ""
    text = v.text or ""
    if cell.get("t") == "s" and text.isdigit() and int(text) < len(shared_strings):
        return shared_strings[int(text)]
    return text


def file_sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


SOURCE_DATE_RANGE_RE = re.compile(
    r"^(\d{4})/(\d{2})/(\d{2})\s*(?:~|～)\s*(?:(\d{4})/(\d{2})/(\d{2})|(\d{2})/(\d{2}))$"
)

# These are display anchors, not a search heuristic.  They were verified
# against the source workbooks' literal Excel-visible period labels.  A moved,
# formula-driven, or internally inconsistent title must stop the run rather
# than silently selecting another date-looking cell.
SOURCE_DATE_ANCHORS = {
    "kpi": {
        "sheet": "上線數KPI_達成率",
        "cells": ("D6", "C10", "C57"),
    },
    "store": {
        "sheet": "上線數KPI_店點達成率_明細",
        "cells": ("H6",),
    },
    "person": {
        "sheet": "手機競賽_個人達成率",
        "cells": ("D6",),
    },
}


def source_data_date_from_range(source_date_range, report_run_date):
    text = str(source_date_range or "").strip()
    match = SOURCE_DATE_RANGE_RE.fullmatch(text)
    if not match:
        raise RuntimeError(f"來源資料區間格式無法驗證：{text}")
    start = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
    if match.group(4):
        end = datetime(int(match.group(4)), int(match.group(5)), int(match.group(6))).date()
    else:
        end = datetime(int(match.group(1)), int(match.group(7)), int(match.group(8))).date()
    if end < start:
        raise RuntimeError(f"來源資料區間起訖逆序：{text}")
    if end > report_run_date:
        raise RuntimeError(f"來源資料截止日 {end.isoformat()} 晚於執行日 {report_run_date.isoformat()}")
    return end.isoformat()


def source_date_provenance_from_xlsx(source, source_kind):
    """Read the source's literal Excel-visible business period at fixed anchors."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("台獎同批來源 gate 需要 bundled Python 的 openpyxl") from exc
    anchor = SOURCE_DATE_ANCHORS[source_kind]
    formula_workbook = load_workbook(source, read_only=True, data_only=False, keep_links=False)
    cached_workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name = anchor["sheet"]
        if sheet_name not in formula_workbook.sheetnames or sheet_name not in cached_workbook.sheetnames:
            raise RuntimeError(f"{source_kind} 來源缺少工作表：{sheet_name}")
        formula_sheet = formula_workbook[sheet_name]
        cached_sheet = cached_workbook[sheet_name]
        cells = []
        for cell_ref in anchor["cells"]:
            formula_cell = formula_sheet[cell_ref]
            cached_cell = cached_sheet[cell_ref]
            formula_value = formula_cell.value
            cached_value = cached_cell.value
            if formula_cell.data_type == "f":
                raise RuntimeError(
                    f"{source_kind} 資料期間不可使用公式：{sheet_name}!{cell_ref}"
                )
            if not isinstance(formula_value, str) or not isinstance(cached_value, str):
                raise RuntimeError(
                    f"{source_kind} 資料期間必須是 literal text：{sheet_name}!{cell_ref}"
                )
            formula_text = formula_value.strip()
            cached_text = cached_value.strip()
            if formula_text != cached_text:
                raise RuntimeError(
                    f"{source_kind} 資料期間公式層/快取層不一致：{sheet_name}!{cell_ref}"
                )
            if not SOURCE_DATE_RANGE_RE.fullmatch(formula_text):
                raise RuntimeError(
                    f"{source_kind} 資料期間格式無法驗證：{sheet_name}!{cell_ref}={formula_text}"
                )
            cells.append({
                "cell": cell_ref,
                "formula": None,
                "formula_layer_value": formula_text,
                "cached_value": cached_text,
                "display_value": formula_text,
                "number_format": formula_cell.number_format,
            })
        ranges = {entry["display_value"] for entry in cells}
        if len(ranges) != 1:
            raise RuntimeError(
                f"{source_kind} 資料期間錨點不一致：{sheet_name} "
                + ", ".join(f"{entry['cell']}={entry['display_value']}" for entry in cells)
            )
        return {
            "source_date_range": cells[0]["display_value"],
            "date_provenance": {
                "sheet": sheet_name,
                "value_kind": "literal-string",
                "cells": cells,
            },
        }
    finally:
        formula_workbook.close()
        cached_workbook.close()


def kpi_identity_for_batch(report_run_date):
    direct_source = Path(os.environ.get("REPORT_KPI_SOURCE", "")).expanduser()
    cloud_identity = None
    if direct_source.is_file():
        kpi_path = direct_source
        cloud_identity = (
            cloud_source_identity("kpi", kpi_path)
            if PHONE_AWARDS_SOURCE_MODE in PHONE_AWARDS_CLOUD_PROVIDERS else None
        )
        kpi_provenance = source_date_provenance_from_xlsx(kpi_path, "kpi")
        kpi_range = kpi_provenance["source_date_range"]
    else:
        try:
            report = json.loads(TODAY_REPORT_DATA_PATH.read_text(encoding="utf-8"))
        except FileNotFoundError as exc:
            raise RuntimeError("同批來源 gate 缺少 today_report_data.json，禁止產出台獎") from exc
        if str(report.get("report_date_iso") or "") != report_run_date.isoformat():
            raise RuntimeError("同批來源 gate KPI report_date_iso 與本次執行日不一致")
        kpi_range = str(report.get("source_date_range") or "").strip()
        kpi_path = Path(str(report.get("source_path") or "")).expanduser()
        if not kpi_path.is_file():
            raise RuntimeError("同批來源 gate KPI source_path 不存在，禁止產出台獎")
        kpi_provenance = source_date_provenance_from_xlsx(kpi_path, "kpi")
        if kpi_range != kpi_provenance["source_date_range"]:
            raise RuntimeError("同批來源 gate KPI parser/Excel 顯示資料期間不一致")
    expected_name = report_run_date.strftime("%m%d.xlsx")
    observed_name = cloud_identity["canonical_basename"] if cloud_identity else kpi_path.name
    if observed_name != expected_name:
        raise RuntimeError(f"同批來源 gate KPI 檔名不一致：預期 {expected_name}，實際 {observed_name}")
    identity = {
        "source_file": kpi_path.name,
        "absolute_path": str(kpi_path.resolve()),
        "source_date_range": kpi_range,
        "source_data_date": source_data_date_from_range(kpi_range, report_run_date),
        "sha256": file_sha256(kpi_path),
        "date_provenance": kpi_provenance["date_provenance"],
    }
    if cloud_identity:
        identity.update(cloud_identity)
        identity["source_file"] = expected_name
        identity["source_date_range"] = kpi_range
        identity["source_data_date"] = source_data_date_from_range(kpi_range, report_run_date)
        identity["date_provenance"] = kpi_provenance["date_provenance"]
    return identity


def validate_same_source_batch(freshness, store_source, person_source):
    """Require KPI and both award sources to carry the same business cutoff before Y26 writes."""
    report_run_date = required_report_run_date()
    kpi = kpi_identity_for_batch(report_run_date)
    kpi_date = kpi["source_data_date"]

    source_specs = {"store": store_source, "person": person_source}
    for kind, source in source_specs.items():
        entry = freshness["source_identity"][kind]
        date_evidence = source_date_provenance_from_xlsx(source, kind)
        source_range = date_evidence["source_date_range"]
        source_date = source_data_date_from_range(source_range, report_run_date)
        if source_date != kpi_date:
            raise RuntimeError(
                f"同批來源 gate KPI/台獎 {kind} 資料截止日不一致："
                f"KPI {kpi_date}，台獎 {source_date}"
            )
        entry["source_date_range"] = source_range
        entry["source_data_date"] = source_date
        entry["date_provenance"] = date_evidence["date_provenance"]
        staged = (
            Path(entry["staged_path"]).expanduser()
            if entry.get("provider") in PHONE_AWARDS_CLOUD_PROVIDERS and entry.get("staged_path")
            else PHONE_INPUT_DIR / entry["canonical_basename"]
        )
        if not staged.is_file():
            raise RuntimeError(f"同批來源 gate 缺少 {kind} staged 檔，禁止產出台獎")
        staged_hash = file_sha256(staged)
        if staged_hash != entry["sha256"]:
            raise RuntimeError(f"同批來源 gate {kind} 原始與 staged SHA-256 不一致")
        entry["staged_path"] = str(staged.resolve())
        entry["staged_sha256"] = staged_hash

    record = {
        "schema_version": "north12b-source-batch/v1",
        "report_run_date": report_run_date.isoformat(),
        "data_cutoff_date": kpi_date,
        "kpi": kpi,
        "awards": freshness["source_identity"],
    }
    canonical = json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    record["batch_id"] = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return record


def load_active_award_config():
    active_path = CONFIG_DIR / "award-config-active.json"
    if os.environ.get("PHONE_AWARD_CONFIG"):
        config_path = Path(os.environ["PHONE_AWARD_CONFIG"]).expanduser()
    elif active_path.exists():
        active = json.loads(active_path.read_text(encoding="utf-8"))
        config_path = CONFIG_DIR / active["activeConfig"]
    else:
        return None
    config = json.loads(config_path.read_text(encoding="utf-8"))
    config["_path"] = str(config_path)
    config["_hash"] = file_sha256(config_path)
    return config


def apply_award_config(config):
    global VISIBLE_STORE_ITEMS, ITEM_ALIASES, MANAGER_AWARD_RULES, SUPERVISOR_AWARD_RULES, MODEL_DISPLAY_NAMES, ACTIVE_AWARD_CONFIG
    if not config:
        return
    selected = config.get("selectedModels") or []
    rules = {row.get("role"): row.get("amounts", {}) for row in config.get("rewardRules", [])}
    if not selected:
        raise ValueError(f"{config.get('_path')} missing selectedModels")
    if set(rules.get("manager", {})) != set(selected) or set(rules.get("supervisor", {})) != set(selected):
        raise ValueError(f"{config.get('_path')} rewardRules must cover every selected model")
    VISIBLE_STORE_ITEMS = list(selected)
    ITEM_ALIASES = {**ITEM_ALIASES, **(config.get("modelAliases") or {})}
    MANAGER_AWARD_RULES = {name: list(values) for name, values in rules["manager"].items()}
    SUPERVISOR_AWARD_RULES = {name: list(values) for name, values in rules["supervisor"].items()}
    MODEL_DISPLAY_NAMES = {
        group.get("modelId"): group.get("displayName")
        for group in config.get("modelGroups", [])
        if group.get("modelId") and group.get("displayName")
    }
    ACTIVE_AWARD_CONFIG = config


def to_num(value):
    if value in ("", None):
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def display_amount(value):
    if value is None or value == "":
        return "—"
    numeric = to_num(value)
    return f"{numeric:,.0f}" if numeric.is_integer() else f"{numeric:,.1f}"


def read_first_sheet(base):
    return read_sheet(base, 0)


def sheet_targets(base):
    wb = ET.parse(base / "xl/workbook.xml").getroot()
    rels = ET.parse(base / "xl/_rels/workbook.xml.rels").getroot()
    relmap = {r.get("Id"): r.get("Target") for r in rels}
    targets = []
    for sheet in wb.find(q("sheets")):
        targets.append((sheet.get("name"), relmap[sheet.get(f"{{{REL_NS}}}id")]))
    return targets


def read_sheet(base, sheet_name_or_index):
    shared = read_shared_strings(base)
    targets = sheet_targets(base)
    if isinstance(sheet_name_or_index, int):
        target = targets[sheet_name_or_index][1]
    else:
        matches = [target for name, target in targets if name == sheet_name_or_index]
        if not matches:
            raise FileNotFoundError(f"Sheet not found in {base}: {sheet_name_or_index}")
        target = matches[0]
    sheet_path = base / "xl" / target.lstrip("/")
    root = ET.parse(sheet_path).getroot()
    cells = {cell.get("r"): cell for cell in root.findall(f".//{q('c')}")}
    return shared, cells


def value(cells, shared, col, row):
    return text_of_cell(cells.get(f"{col}{row}"), shared)


def formula(cells, col, row):
    cell = cells.get(f"{col}{row}")
    if cell is None:
        return ""
    formula_el = cell.find(q("f"))
    return formula_el.text or "" if formula_el is not None else ""


def award_values_from_formula(y26_base, formula_text, expected_sheet):
    match = re.fullmatch(r"'?([^']+)'?!\$?([A-Z]+)\$?(\d+)", formula_text.strip())
    if not match or match.group(1) != expected_sheet:
        raise ValueError(f"Invalid {expected_sheet} award formula: {formula_text!r}")
    row = int(match.group(3))
    shared, cells = read_sheet(y26_base, expected_sheet)
    values = [to_num(value(cells, shared, num_to_col(col), row)) for col in range(col_to_num("C"), col_to_num("G") + 1)]
    if len(values) != 5 or any(amount <= 0 for amount in values):
        raise ValueError(f"Invalid {expected_sheet} award values at row {row}: {values}")
    return [int(amount) if amount.is_integer() else amount for amount in values]


def load_award_rules_from_y26(y26_base, groups):
    shared, cells = read_sheet(y26_base, "台獎機款")
    manager_rules = {}
    supervisor_rules = {}
    group_index = 0
    for block in AWARD_RULE_BLOCKS:
        for start_col in block["starts"]:
            if group_index >= len(groups):
                break
            group_name = groups[group_index]["name"]
            # Y26 reuses campaign slots while display names change, so the formula
            # position is authoritative for the current tracked-item ordering.
            manager_rules[group_name] = award_values_from_formula(
                y26_base,
                formula(cells, start_col, block["manager_row"]),
                "店",
            )
            supervisor_rules[group_name] = award_values_from_formula(
                y26_base,
                formula(cells, start_col, block["supervisor_row"]),
                "督",
            )
            group_index += 1
    expected = {group["name"] for group in groups}
    if set(manager_rules) != expected or set(supervisor_rules) != expected:
        raise ValueError(
            f"Incomplete Y26 award rules: manager={len(manager_rules)}, "
            f"supervisor={len(supervisor_rules)}, expected={len(expected)}"
        )
    return manager_rules, supervisor_rules


def load_central_award_rules(groups):
    expected = set(VISIBLE_STORE_ITEMS)
    actual = {group["name"] for group in groups}
    if actual != expected or len(groups) != 13:
        missing = sorted(expected - actual)
        unexpected = sorted(actual - expected)
        raise ValueError(
            f"Expected 13 central phone-award items, got {len(groups)}; "
            f"missing={missing}, unexpected={unexpected}"
        )
    return deepcopy(MANAGER_AWARD_RULES), deepcopy(SUPERVISOR_AWARD_RULES)


def item_groups(cells, shared, header_row, subheader_row, start_col, max_col):
    groups = []
    seen = set()
    col = start_col
    while col <= max_col:
        name = value(cells, shared, num_to_col(col), header_row)
        if name.startswith("上線數_") or name.startswith("空機數_"):
            actual_header = value(cells, shared, num_to_col(col), subheader_row)
            target_offset = None
            rate_offset = None
            for offset in range(1, 5):
                header = value(cells, shared, num_to_col(col + offset), subheader_row)
                if target_offset is None and "目標數" in header:
                    target_offset = offset
                if "達成率" in header:
                    rate_offset = offset
                    break
            if "實際數" in actual_header and target_offset is not None and rate_offset is not None:
                short = re.sub(r"^(上線數|空機數)_\d+_", "", name)
                short = ITEM_ALIASES.get(short, short)
                if short not in seen:
                    seen.add(short)
                    groups.append({
                        "name": short,
                        "actual_col": num_to_col(col),
                        "target_col": num_to_col(col + target_offset),
                        "rate_col": num_to_col(col + rate_offset),
                    })
        col += 1
    return groups


def normalized_header(text):
    return re.sub(r"\s+", "", (text or "").replace("\u3000", ""))


def summary_columns(cells, shared, subheader_rows, start_col, max_col):
    rows = [subheader_rows] if isinstance(subheader_rows, int) else list(subheader_rows)
    columns = {}
    for row in rows:
        for col in range(start_col, max_col + 1):
            header = normalized_header(value(cells, shared, num_to_col(col), row))
            if not header:
                continue
            if "actual_total" not in columns and ("實際總獎金" in header or header == "實際獎金"):
                columns["actual_total"] = num_to_col(col)
            elif "projected" not in columns and ("推估原始總獎金" in header or "推估獎金" in header or "預估獎金" in header):
                columns["projected"] = num_to_col(col)
            elif "rank" not in columns and (header == "排名" or header == "名次" or header.endswith("排名")):
                columns["rank"] = num_to_col(col)
            elif "award" not in columns and ("領獎與否" in header or "是否領獎" in header or header == "領獎"):
                columns["award"] = num_to_col(col)
    return columns


def visible_groups(group_map):
    groups = []
    for name in VISIBLE_STORE_ITEMS:
        if name in group_map:
            groups.append(group_map[name])
    return groups


def group_values(cells, shared, group, row):
    if "components" not in group:
        actual = to_num(value(cells, shared, group["actual_col"], row))
        target = to_num(value(cells, shared, group["target_col"], row))
        source_rate = to_num(value(cells, shared, group["rate_col"], row))
        actual_rate = actual / target if target > 0 else 0
        return {
            "name": group["name"],
            "actual": actual,
            "target": target,
            "rate": source_rate,
            "actual_rate": actual_rate,
        }

    actual = 0.0
    target = 0.0
    for component in group["components"]:
        actual += to_num(value(cells, shared, component["actual_col"], row))
        target += to_num(value(cells, shared, component["target_col"], row))
    actual_rate = actual / target if target > 0 else 0
    return {
        "name": group["name"],
        "actual": actual,
        "target": target,
        "rate": actual_rate,
        "actual_rate": actual_rate,
    }


def build_store_awards(source="auto"):
    if source == "y26" and Y26_SRC.exists():
        return build_store_awards_from_y26()
    return build_store_awards_from_source()


def build_store_awards_from_source():
    shared, cells = read_first_sheet(STORE_SRC)
    right_col = max_cell_col(cells, col_to_num("JS"))
    all_groups = item_groups(cells, shared, 10, 11, col_to_num("K"), right_col)
    group_map = {group["name"]: group for group in all_groups}
    groups = visible_groups(group_map)
    totals = summary_columns(cells, shared, [11, 16], col_to_num("K"), right_col)
    rows = []
    for row in [12] + list(range(17, 26)):
        store = value(cells, shared, "I", row)
        if not store:
            store = value(cells, shared, "F", row)
        if not store or store == "北一二B":
            store = "北一二B整體"
        code = value(cells, shared, "G", row) if row != 12 else ""
        row_out = {
            "store": store,
            "code": code,
            "items": [],
            "actual_total": to_num(value(cells, shared, totals["actual_total"], row)) if "actual_total" in totals else None,
            "projected": to_num(value(cells, shared, totals["projected"], row)) if "projected" in totals else None,
            "rank": value(cells, shared, totals["rank"], row) if "rank" in totals else "",
            "award": value(cells, shared, totals["award"], row).strip() if "award" in totals else "",
        }
        for group in groups:
            row_out["items"].append(group_values(cells, shared, group, row))
        rows.append(row_out)
    return groups, rows


def build_person_awards(source="auto"):
    if source == "y26" and Y26_SRC.exists():
        return build_person_awards_from_y26()
    return build_person_awards_from_source()


def build_person_awards_from_source():
    shared, cells = read_first_sheet(PERSON_SRC)
    right_col = max_cell_col(cells, col_to_num("JQ"))
    groups = item_groups(cells, shared, 9, 10, col_to_num("I"), right_col)
    totals = summary_columns(cells, shared, 10, col_to_num("I"), right_col)
    people = []
    for row in range(11, 52):
        actual_total = to_num(value(cells, shared, totals["actual_total"], row)) if "actual_total" in totals else 0
        projected = to_num(value(cells, shared, totals["projected"], row)) if "projected" in totals else None
        award = value(cells, shared, totals["award"], row).strip() if "award" in totals else ("Y" if actual_total > 0 else "N")
        rank_value = int(to_num(value(cells, shared, totals["rank"], row))) if "rank" in totals else 0
        people.append({
            "store": value(cells, shared, "D", row),
            "title": value(cells, shared, "E", row),
            "emp_id": value(cells, shared, "F", row),
            "name": value(cells, shared, "G", row),
            "actual_total": actual_total,
            "projected": projected,
            "rank": rank_value,
            "award": award,
            "units": sum(to_num(value(cells, shared, g["actual_col"], row)) for g in groups),
        })
    people.sort(key=lambda r: (r["award"] != "Y", r["rank"] == 0, r["rank"] if r["rank"] > 0 else 999999, -r["actual_total"], r["store"], r["emp_id"]))
    return people


def set_cell_value(cells, ref, new_value):
    cell = cells.get(ref)
    if cell is None:
        return
    for child in list(cell):
        if child.tag in {q("f"), q("v"), q("is")}:
            cell.remove(child)
    if new_value is None or new_value == "":
        cell.attrib.pop("t", None)
        return
    if isinstance(new_value, (int, float)) and not isinstance(new_value, bool) and math.isfinite(new_value):
        cell.attrib.pop("t", None)
        ET.SubElement(cell, q("v")).text = str(int(new_value) if float(new_value).is_integer() else round(new_value, 6))
        return
    cell.set("t", "inlineStr")
    is_el = ET.SubElement(cell, q("is"))
    ET.SubElement(is_el, q("t")).text = str(new_value)


def sheet_path_for(base, sheet_name):
    targets = sheet_targets(base)
    matches = [target for name, target in targets if name == sheet_name]
    if not matches:
        raise FileNotFoundError(f"Sheet not found in {base}: {sheet_name}")
    return base / "xl" / matches[0].lstrip("/")


def shift_row_ref(ref, offset):
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    if not match:
        return ref
    return f"{match.group(1)}{int(match.group(2)) + offset}"


def clone_y26_block(root, source_start, source_end, target_start):
    """Extend the legacy Y26 sheet with another block without changing its cell styling."""
    sheet_data = root.find(q("sheetData"))
    if sheet_data is None:
        raise ValueError("Y26 台獎機款 sheet has no sheetData")
    target_end = target_start + (source_end - source_start)
    for row in list(sheet_data.findall(q("row"))):
        row_number = int(row.get("r"))
        if target_start <= row_number <= target_end:
            sheet_data.remove(row)
    source_rows = {
        int(row.get("r")): row
        for row in sheet_data.findall(q("row"))
        if source_start <= int(row.get("r")) <= source_end
    }
    for row_number in range(source_start, source_end + 1):
        source_row = source_rows.get(row_number)
        if source_row is None:
            continue
        clone = deepcopy(source_row)
        clone.set("r", str(row_number + target_start - source_start))
        for cell in clone.findall(q("c")):
            cell.set("r", shift_row_ref(cell.get("r"), target_start - source_start))
        sheet_data.append(clone)

    merge_cells = root.find(q("mergeCells"))
    if merge_cells is None:
        merge_cells = ET.SubElement(root, q("mergeCells"))
    existing_refs = {merge.get("ref") for merge in merge_cells.findall(q("mergeCell"))}
    for merge in list(merge_cells.findall(q("mergeCell"))):
        ref = merge.get("ref", "")
        row_numbers = [int(value) for value in re.findall(r"\d+", ref)]
        if row_numbers and min(row_numbers) >= source_start and max(row_numbers) <= source_end:
            shifted = ":".join(shift_row_ref(part, target_start - source_start) for part in ref.split(":"))
            if shifted not in existing_refs:
                ET.SubElement(merge_cells, q("mergeCell"), {"ref": shifted})
                existing_refs.add(shifted)
    merge_cells.set("count", str(len(merge_cells.findall(q("mergeCell")))) )
    dimension = root.find(q("dimension"))
    if dimension is not None:
        current_ref = dimension.get("ref", "A1:AD1")
        start_ref, end_ref = current_ref.split(":", 1)
        end_match = re.fullmatch(r"([A-Z]+)(\d+)", end_ref)
        if end_match and int(end_match.group(2)) < target_end:
            dimension.set("ref", f"{start_ref}:{end_match.group(1)}{target_end}")


def update_y26_store_sheet(y26_base, groups, data):
    sheet_path = sheet_path_for(y26_base, "台獎機款")
    root = ET.parse(sheet_path).getroot()
    starts = ["F", "J", "N", "R", "V", "Z"]
    if len(groups) > 12:
        clone_y26_block(root, 17, 32, 52)
    cells = {cell.get("r"): cell for cell in root.findall(f".//{q('c')}")}
    blocks = [
        {"group_offset": 0, "header_row": 1, "rows": list(range(7, 17)), "starts": starts},
        {"group_offset": 6, "header_row": 17, "rows": list(range(23, 33)), "starts": starts},
    ]
    if len(groups) > 12:
        blocks.append({"group_offset": 12, "header_row": 52, "rows": list(range(58, 68)), "starts": starts})
    for block in blocks:
        for idx, start_col in enumerate(block["starts"], start=block["group_offset"]):
            set_cell_value(
                cells,
                f"{start_col}{block['header_row']}",
                display_item_name(groups[idx]["name"]) if idx < len(groups) else "",
            )
        for target_row, row_data in zip(block["rows"], data):
            store = "北一二B" if row_data["store"] == "北一二B整體" else row_data["store"]
            for col, val in [
                ("A", store),
                ("B", row_data["actual_total"]),
                ("C", row_data["projected"]),
                ("D", row_data["rank"]),
                ("E", row_data["award"]),
            ]:
                set_cell_value(cells, f"{col}{target_row}", val)
            for idx, start_col in enumerate(block["starts"], start=block["group_offset"]):
                item = row_data["items"][idx] if idx < len(row_data["items"]) and idx < len(groups) else None
                actual_col_num = col_to_num(start_col)
                set_cell_value(cells, f"{start_col}{target_row}", item["actual"] if item else "")
                set_cell_value(cells, f"{num_to_col(actual_col_num + 1)}{target_row}", item["target"] if item else "")
                set_cell_value(cells, f"{num_to_col(actual_col_num + 2)}{target_row}", item["actual"] - item["target"] if item else "")
                set_cell_value(cells, f"{num_to_col(actual_col_num + 3)}{target_row}", item["rate"] if item else "")
    ET.ElementTree(root).write(sheet_path, encoding="utf-8", xml_declaration=True)


def update_y26_person_sheet(y26_base, people):
    sheet_path = sheet_path_for(y26_base, "個人台獎")
    root = ET.parse(sheet_path).getroot()
    cells = {cell.get("r"): cell for cell in root.findall(f".//{q('c')}")}
    start_row = 2
    for idx, person in enumerate(people, start=start_row):
        for col, val in [
            ("A", "北一二B"),
            ("B", person["store"]),
            ("C", person["title"]),
            ("D", person["emp_id"]),
            ("E", person["name"]),
            ("F", person["units"]),
            ("G", person["actual_total"]),
            ("H", person["projected"]),
            ("I", person["rank"]),
            ("J", person["award"]),
            ("K", "未領獎" if person["award"] != "Y" else ""),
        ]:
            set_cell_value(cells, f"{col}{idx}", val)
    for idx in range(start_row + len(people), 80):
        for col in "ABCDEFGHIJK":
            set_cell_value(cells, f"{col}{idx}", "")
    ET.ElementTree(root).write(sheet_path, encoding="utf-8", xml_declaration=True)


def update_y26_workbook(y26_base, groups, data, people):
    update_y26_store_sheet(y26_base, groups, data)
    update_y26_person_sheet(y26_base, people)


def build_store_awards_from_y26():
    shared, cells = read_sheet(Y26_SRC, "台獎機款")
    item_blocks = [
        {"header_row": 1, "data_rows": range(7, 17), "starts": ["F", "J", "N", "R", "V"]},
        {"header_row": 17, "data_rows": range(23, 33), "starts": ["F", "J", "N", "R", "V"]},
    ]
    groups = []
    row_map = {}

    for block in item_blocks:
        block_groups = []
        for start_col in block["starts"]:
            name = value(cells, shared, start_col, block["header_row"]).replace("\u00a0", " ").strip()
            if not name:
                continue
            group = {
                "name": name,
                "actual_col": start_col,
                "target_col": num_to_col(col_to_num(start_col) + 1),
                "rate_col": num_to_col(col_to_num(start_col) + 3),
            }
            groups.append(group)
            block_groups.append(group)

        for row in block["data_rows"]:
            store = value(cells, shared, "A", row).strip()
            if not store:
                continue
            if store == "北一二B":
                store = "北一二B整體"
            row_out = row_map.setdefault(store, {
                "store": store,
                "code": "",
                "items": [],
                "actual_total": to_num(value(cells, shared, "B", row)),
                "projected": to_num(value(cells, shared, "C", row)),
                "rank": value(cells, shared, "D", row),
                "award": value(cells, shared, "E", row).strip() or "N",
            })
            for group in block_groups:
                actual = to_num(value(cells, shared, group["actual_col"], row))
                target = to_num(value(cells, shared, group["target_col"], row))
                source_rate = to_num(value(cells, shared, group["rate_col"], row))
                actual_rate = actual / target if target > 0 else 0
                row_out["items"].append({"name": group["name"], "actual": actual, "target": target, "rate": source_rate, "actual_rate": actual_rate})

    return groups, list(row_map.values())


def build_person_awards_from_y26():
    shared, cells = read_sheet(Y26_SRC, "個人台獎")
    people = []
    for row in range(2, 80):
        store = value(cells, shared, "B", row).strip()
        award = value(cells, shared, "J", row).strip()
        name = value(cells, shared, "E", row).strip()
        if not store or store == "0" or award not in {"Y", "N"} or not name or name == "0":
            continue
        people.append({
            "store": store,
            "title": value(cells, shared, "C", row),
            "emp_id": value(cells, shared, "D", row),
            "name": name,
            "actual_total": to_num(value(cells, shared, "G", row)),
            "projected": to_num(value(cells, shared, "H", row)),
            "rank": int(to_num(value(cells, shared, "I", row))),
            "award": award,
            "units": to_num(value(cells, shared, "F", row)),
        })
    people.sort(key=lambda r: (r["award"] != "Y", r["rank"] == 0, r["rank"] if r["rank"] > 0 else 999999, -r["actual_total"], r["store"], r["emp_id"]))
    return people


def copy_tree(src, dst):
    if dst.exists():
        shutil.rmtree(dst)
    shutil.copytree(src, dst)


def newest_existing(paths):
    existing = [path for path in paths if path.exists()]
    if not existing:
        return None
    return max(existing, key=lambda path: path.stat().st_mtime)


AWARD_SOURCE_BASENAMES = {
    "store": "01-08-03-(密)直營_手機競賽日報_店點達成率、排名及獎金.xlsx",
    "person": "01-08-04-(密)直營_手機競賽日報_個人達成率、排名及獎金.xlsx",
}
TAIPEI = timezone(timedelta(hours=8))


def required_report_run_date():
    """Daily award artifacts have no safe date fallback: the runner must name its run."""
    try:
        return datetime.strptime(REPORT_RUN_DATE, "%Y-%m-%d").date()
    except ValueError as exc:
        raise RuntimeError("REPORT_DATE_ISO is required for 台獎來源 freshness gate") from exc


def canonical_award_source_basename(path, kind):
    """Normalize only Finder's duplicate counter, never infer a date from a filename."""
    raw = Path(path).name
    expected = AWARD_SOURCE_BASENAMES[kind]
    stem, suffix = expected[:-5], expected[-5:]
    duplicate = re.fullmatch(re.escape(stem) + r"\s+\d+" + re.escape(suffix), raw)
    canonical = expected if duplicate else raw
    if canonical != expected:
        raise ValueError(
            f"{kind} source canonical basename mismatch: expected {expected}, got {raw}"
        )
    return canonical


def path_is_inside(path, directory):
    try:
        Path(path).resolve().relative_to(Path(directory).resolve())
        return True
    except ValueError:
        return False


def load_cloud_source_manifest():
    global CLOUD_SOURCE_MANIFEST
    if CLOUD_SOURCE_MANIFEST is not None:
        return CLOUD_SOURCE_MANIFEST
    if PHONE_AWARDS_SOURCE_MODE not in PHONE_AWARDS_CLOUD_PROVIDERS:
        raise RuntimeError("cloud source manifest requested outside a supported cloud provider")
    if not PHONE_AWARDS_CLOUD_MANIFEST_PATH:
        raise RuntimeError(f"{PHONE_AWARDS_SOURCE_MODE} source manifest is required; fallback is OFF")
    manifest_path = Path(PHONE_AWARDS_CLOUD_MANIFEST_PATH).expanduser()
    if not manifest_path.is_file():
        raise FileNotFoundError(f"cloud source manifest not found: {manifest_path}")
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    expected_schema = {
        "onedrive-cloud": "north12b-onedrive-cloud-source/v1",
        "google-drive-cloud": "north12b-google-drive-cloud-source/v1",
    }[PHONE_AWARDS_SOURCE_MODE]
    if payload.get("provider") != PHONE_AWARDS_SOURCE_MODE:
        raise RuntimeError("cloud source manifest provider mismatch")
    if payload.get("schema_version") != expected_schema:
        raise RuntimeError("cloud source manifest schema is invalid")
    if str(payload.get("report_run_date") or "") != REPORT_RUN_DATE:
        raise RuntimeError("cloud source manifest report_run_date mismatch")
    run_id = str(payload.get("run_id") or "").strip()
    if not run_id:
        raise RuntimeError("cloud source manifest run_id is missing")
    if os.environ.get("REPORT_RUN_ID", "").strip() not in {"", run_id}:
        raise RuntimeError("cloud source manifest run_id mismatch")
    if not isinstance(payload.get("sources"), dict):
        raise RuntimeError("cloud source manifest sources are missing")
    CLOUD_SOURCE_MANIFEST = payload
    return payload


def cloud_source_identity(kind, source):
    manifest = load_cloud_source_manifest()
    entry = manifest["sources"].get(kind)
    if not isinstance(entry, dict):
        raise RuntimeError(f"OneDrive cloud source manifest missing {kind}")
    expected = (
        required_report_run_date().strftime("%m%d.xlsx")
        if kind == "kpi" else AWARD_SOURCE_BASENAMES[kind]
    )
    canonical = str(entry.get("canonical_basename") or entry.get("name") or "")
    if canonical != expected:
        raise RuntimeError(f"{kind} cloud canonical basename mismatch: expected {expected}, got {canonical}")
    provider = str(entry.get("provider") or "")
    drive_item_id = str(entry.get("driveItemId") or "").strip()
    modified = str(entry.get("lastModifiedDateTime") or "").strip()
    etag = str(entry.get("eTag") or "").strip()
    run_id = str(entry.get("run_id") or "").strip()
    size = entry.get("size")
    digest = str(entry.get("sha256") or "").lower()
    staged_digest = str(entry.get("staged_sha256") or "").lower()
    staged_path = Path(str(entry.get("staged_path") or "")).expanduser()
    if provider != PHONE_AWARDS_SOURCE_MODE or not drive_item_id or not modified:
        raise RuntimeError(f"{kind} {PHONE_AWARDS_SOURCE_MODE} version identity is incomplete")
    if provider == "onedrive-cloud" and not etag:
        raise RuntimeError(f"{kind} OneDrive eTag is missing")
    try:
        datetime.fromisoformat(modified.replace("Z", "+00:00"))
    except ValueError as exc:
        raise RuntimeError(f"{kind} cloud lastModifiedDateTime is invalid") from exc
    if not isinstance(size, int) or size <= 0:
        raise RuntimeError(f"{kind} cloud size is invalid")
    if not re.fullmatch(r"[a-f0-9]{64}", digest) or staged_digest != digest:
        raise RuntimeError(f"{kind} cloud/staged SHA-256 is invalid")
    if run_id != str(manifest.get("run_id") or ""):
        raise RuntimeError(f"{kind} cloud run_id mismatch")
    if staged_path.resolve() != Path(source).resolve():
        raise RuntimeError(f"{kind} staged path does not match cloud manifest")
    if not staged_path.is_file() or staged_path.stat().st_size != size:
        raise RuntimeError(f"{kind} staged file size does not match cloud metadata")
    actual = file_sha256(staged_path)
    if actual != digest:
        raise RuntimeError(f"{kind} staged SHA-256 does not match cloud download")
    identity = {
        "provider": provider,
        "driveItemId": drive_item_id,
        "basename": canonical,
        "canonical_basename": canonical,
        "lastModifiedDateTime": modified,
        "size": size,
        "sha256": digest,
        "run_id": run_id,
        "origin": provider,
        "absolute_path": str(staged_path.resolve()),
        "staged_path": str(staged_path.resolve()),
        "staged_sha256": staged_digest,
    }
    if etag:
        identity["eTag"] = etag
    if provider == "google-drive-cloud":
        identity["googleDriveFileId"] = drive_item_id
    return identity


def resolve_exact_or_fresh_source(env_path, prefix, kind, expected_data_date):
    """Production uses an exact cloud manifest; local scanning is emergency-only."""
    if PHONE_AWARDS_SOURCE_MODE in PHONE_AWARDS_CLOUD_PROVIDERS:
        if not env_path:
            raise RuntimeError(f"{kind} {PHONE_AWARDS_SOURCE_MODE} staged source is required; fallback is OFF")
        exact = Path(env_path).expanduser().resolve()
        if not exact.is_file():
            raise FileNotFoundError(f"exact cloud source not found: {exact}")
        cloud_source_identity(kind, exact)
        return exact, PHONE_AWARDS_SOURCE_MODE
    if PHONE_AWARDS_SOURCE_MODE != "local-emergency" or not PHONE_AWARDS_LOCAL_EMERGENCY_ENABLED:
        raise RuntimeError(
            "production source mode must be google-drive-cloud; local emergency fallback requires "
            "PHONE_AWARDS_SOURCE_MODE=local-emergency and PHONE_AWARDS_LOCAL_EMERGENCY_ENABLED=1"
        )
    if env_path:
        exact = Path(env_path).expanduser().resolve()
        if not exact.is_file():
            raise FileNotFoundError(f"exact source not found: {exact}")
        if path_is_inside(exact, PHONE_INPUT_DIR):
            raise ValueError(f"{kind} source points at canonical staging, not an immutable input: {exact}")
        canonical_award_source_basename(exact, kind)
        return exact, "exact-materialized"

    if not PHONE_AWARDS_ORIGIN_DIR.is_dir():
        raise FileNotFoundError(f"台獎原始來源目錄不存在：{PHONE_AWARDS_ORIGIN_DIR}")
    candidates = [
        path for path in PHONE_AWARDS_ORIGIN_DIR.glob(f"{prefix}*.xlsx")
        if path.is_file()
    ]
    if not candidates:
        raise FileNotFoundError(f"今日台獎來源尚未更新：缺少 {AWARD_SOURCE_BASENAMES[kind]}")
    report_run_date = required_report_run_date()
    verified = []
    diagnostics = []
    for candidate in candidates:
        try:
            canonical_award_source_basename(candidate, kind)
            evidence = source_date_provenance_from_xlsx(candidate, kind)
            cutoff = source_data_date_from_range(evidence["source_date_range"], report_run_date)
            verified.append((candidate, cutoff, file_sha256(candidate)))
            diagnostics.append(f"{candidate.name}={cutoff}")
        except Exception as exc:
            diagnostics.append(f"{candidate.name}=unverifiable({exc})")

    matches = [entry for entry in verified if entry[1] == expected_data_date]
    if not matches:
        details = "; ".join(diagnostics) or "no verifiable candidates"
        raise RuntimeError(
            f"今日台獎來源沒有與 KPI {expected_data_date} 對齊的 {kind} 檔：{details}"
        )
    distinct_hashes = {entry[2] for entry in matches}
    if len(distinct_hashes) != 1:
        raise RuntimeError(
            f"今日台獎來源有多個 {kind} 同資料日但 SHA-256 不同的版本；"
            "需提供既有受控來源版本，禁止以 mtime 任選"
        )
    # A matching hash may exist under Finder's duplicate names.  mtime is only
    # an audit/tie-break attribute after identical content and business date
    # have been proven; it is never the selection criterion for a data date.
    source = max((entry[0] for entry in matches), key=lambda path: path.stat().st_mtime)
    return source.resolve(), "onedrive-original-business-date"


def previous_award_source_identity(report_run_date):
    """Find the latest prior formal pair, preferring immutable cloud version identity."""
    if not PHONE_AWARDS_LOG_DIR.is_dir():
        return None
    candidates = []
    for manifest_path in PHONE_AWARDS_LOG_DIR.glob("run-manifest-*.json"):
        match = re.fullmatch(r"run-manifest-(\d{8})\.json", manifest_path.name)
        if not match:
            continue
        try:
            day = datetime.strptime(match.group(1), "%Y%m%d").date()
            payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (ValueError, json.JSONDecodeError):
            continue
        hashes = payload.get("sourceHashes") or {}
        identity = (
            payload.get("awardSourceIdentity")
            or payload.get("sourceIdentity")
            or (payload.get("sourceBatch") or {}).get("awards")
            or {}
        )
        has_pair = all(isinstance(identity.get(kind), dict) for kind in ("store", "person"))
        has_hash_pair = hashes.get("store") and hashes.get("person")
        if day < report_run_date and (has_pair or has_hash_pair):
            candidates.append((day, payload, hashes, identity))
    if not candidates:
        return None
    day, payload, hashes, identity = max(candidates, key=lambda item: item[0])
    return {
        "report_run_date": day.isoformat(),
        "run_id": str(payload.get("runId") or ""),
        "store": identity.get("store") or {"sha256": str(hashes.get("store") or "")},
        "person": identity.get("person") or {"sha256": str(hashes.get("person") or "")},
    }


def unchanged_source_exception():
    """A reused SHA needs a human-verifiable, hash-bound exception; a filename never qualifies."""
    reason = os.environ.get("REPORT_AWARDS_ALLOW_UNCHANGED_REASON", "").strip()
    evidence_path = os.environ.get("REPORT_AWARDS_ALLOW_UNCHANGED_EVIDENCE_PATH", "").strip()
    evidence_hash = os.environ.get("REPORT_AWARDS_ALLOW_UNCHANGED_EVIDENCE_SHA256", "").strip().lower()
    if not any((reason, evidence_path, evidence_hash)):
        return None
    if not reason or not evidence_path or not re.fullmatch(r"[a-f0-9]{64}", evidence_hash):
        raise RuntimeError("台獎來源 SHA 未變更時，必須提供 reason、evidence path 與 SHA-256")
    evidence = Path(evidence_path).expanduser()
    if not evidence.is_file() or file_sha256(evidence) != evidence_hash:
        raise RuntimeError("台獎來源 SHA 未變更例外的 evidence SHA-256 無法驗證")
    return {"reason": reason, "evidence_path": str(evidence.resolve()), "evidence_sha256": evidence_hash}


def validate_fresh_award_sources(store_source, person_source, store_origin, person_origin):
    report_run_date = required_report_run_date()
    source_hashes = {"store": file_sha256(store_source), "person": file_sha256(person_source)}
    run_id = (
        str(load_cloud_source_manifest().get("run_id") or "").strip()
        if PHONE_AWARDS_SOURCE_MODE in PHONE_AWARDS_CLOUD_PROVIDERS else os.environ.get("REPORT_RUN_ID", "").strip()
    ) or f"awards-{report_run_date.strftime('%Y%m%d')}-{source_hashes['store'][:12]}-{source_hashes['person'][:12]}"
    identities = {}
    for kind, source, origin in (
        ("store", store_source, store_origin),
        ("person", person_source, person_origin),
    ):
        if PHONE_AWARDS_SOURCE_MODE in PHONE_AWARDS_CLOUD_PROVIDERS:
            identities[kind] = cloud_source_identity(kind, source)
        else:
            stat = source.stat()
            mtime = datetime.fromtimestamp(stat.st_mtime, TAIPEI)
            identities[kind] = {
                "basename": source.name,
                "canonical_basename": canonical_award_source_basename(source, kind),
                "absolute_path": str(source.resolve()),
                "mtime": mtime.isoformat(timespec="seconds"),
                "size": stat.st_size,
                "sha256": source_hashes[kind],
                "run_id": run_id,
                "origin": origin,
            }

    previous = previous_award_source_identity(report_run_date)
    exception = None
    if PHONE_AWARDS_SOURCE_MODE == "onedrive-cloud" and previous:
        changed = {}
        for kind in ("store", "person"):
            prior = previous[kind]
            current = identities[kind]
            if prior.get("provider") != "onedrive-cloud":
                changed[kind] = True
            else:
                changed[kind] = any(
                    str(current.get(field) or "") != str(prior.get(field) or "")
                    for field in ("driveItemId", "eTag", "lastModifiedDateTime")
                )
        if not changed["store"] and not changed["person"]:
            raise RuntimeError(
                "今日台獎來源尚未更新：store/person cloud item ID、eTag、lastModifiedDateTime 均未變；"
                "不得 fallback 到 local/staging/outputs/cache"
            )
        if not changed["store"] or not changed["person"]:
            raise RuntimeError(
                f"今日台獎來源只更新一份：store={changed['store']}，person={changed['person']}；必須兩份同時更新"
            )
    elif previous:
        unchanged = {
            kind: str(previous[kind].get("sha256") or "") == source_hashes[kind]
            for kind in ("store", "person")
        }
        if unchanged["store"] and unchanged["person"]:
            exception = unchanged_source_exception()
            if not exception:
                raise RuntimeError(
                    "今日台獎來源尚未更新：store/person SHA-256 與前次正式 run 完全相同；"
                    "不得以 cloud modifiedTime、staging mtime、檔名或舊 outputs 發布"
                )
        elif unchanged["store"] or unchanged["person"]:
            raise RuntimeError(
                f"今日台獎來源只更新一份：store={not unchanged['store']}，person={not unchanged['person']}；"
                "必須兩份 raw SHA-256 同時更新"
            )
    return {
        "run_id": run_id,
        "report_run_date": report_run_date.isoformat(),
        "status": "fresh" if not exception else "unchanged-source-verified-exception",
        "previous": previous,
        "unchanged_exception": exception,
        "source_identity": identities,
    }


def resolve_y26_source():
    drive_matches = []
    if PHONE_INPUT_DIR.exists():
        drive_matches.extend(PHONE_INPUT_DIR.glob("Y26*.xlsx"))
        drive_matches.extend(path for path in PHONE_INPUT_DIR.glob("Y26*") if path.is_dir())
    return newest_existing(drive_matches) or Y26_FALLBACK_SRC


def unpack_or_copy_workbook(src, dst):
    if src is None or not src.exists():
        raise FileNotFoundError(f"source not found: {src}")
    if dst.exists():
        shutil.rmtree(dst)
    if src.is_dir():
        shutil.copytree(src, dst)
        return dst
    dst.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(src) as zf:
        zf.extractall(dst)
    return dst


def ensure_styles(base):
    styles_path = base / "xl/styles.xml"
    root = ET.parse(styles_path).getroot()
    fills = root.find(q("fills"))
    cell_xfs = root.find(q("cellXfs"))
    base_xf = deepcopy(cell_xfs[0])

    def add_fill(rgb):
        idx = len(fills)
        fill = ET.SubElement(fills, q("fill"))
        pattern = ET.SubElement(fill, q("patternFill"), {"patternType": "solid"})
        ET.SubElement(pattern, q("fgColor"), {"rgb": rgb})
        ET.SubElement(pattern, q("bgColor"), {"indexed": "64"})
        fills.set("count", str(len(fills)))
        return idx

    def add_xf(fill_id, bold=False):
        idx = len(cell_xfs)
        xf = deepcopy(base_xf)
        xf.set("fillId", str(fill_id))
        xf.set("applyFill", "1")
        if bold:
            xf.set("applyFont", "1")
        cell_xfs.append(xf)
        cell_xfs.set("count", str(len(cell_xfs)))
        return idx

    styles = {
        "header": add_xf(add_fill("FFFFC000")),
        "subheader": add_xf(add_fill("FFD9EAF7")),
        "blue": add_xf(add_fill("FF4EA5F5")),
        "green": add_xf(add_fill("FF00B050")),
        "pink": add_xf(add_fill("FFFFC7CE")),
    }
    ET.ElementTree(root).write(styles_path, encoding="utf-8", xml_declaration=True)
    return styles


def make_cell(ref, value, style=None):
    cell = ET.Element(q("c"), {"r": ref})
    if style is not None:
        cell.set("s", str(style))
    if value is None or value == "":
        return cell
    if isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value):
        ET.SubElement(cell, q("v")).text = str(int(value) if float(value).is_integer() else round(value, 4))
    else:
        cell.set("t", "inlineStr")
        is_el = ET.SubElement(cell, q("is"))
        ET.SubElement(is_el, q("t")).text = str(value)
    return cell


def build_sheet(rows, widths=None, freeze_ref=None):
    root = ET.Element(q("worksheet"))
    if widths:
        cols = ET.SubElement(root, q("cols"))
        for idx, width in enumerate(widths, start=1):
            ET.SubElement(cols, q("col"), {"min": str(idx), "max": str(idx), "width": str(width), "customWidth": "1"})
    if freeze_ref:
        sheet_views = ET.SubElement(root, q("sheetViews"))
        view = ET.SubElement(sheet_views, q("sheetView"), {"workbookViewId": "0"})
        ET.SubElement(view, q("pane"), {"ySplit": str(int(re.search(r"\d+", freeze_ref).group()) - 1), "topLeftCell": freeze_ref, "activePane": "bottomLeft", "state": "frozen"})
    sheet_data = ET.SubElement(root, q("sheetData"))
    for r_idx, row in enumerate(rows, start=1):
        row_el = ET.SubElement(sheet_data, q("row"), {"r": str(r_idx)})
        for c_idx, (val, style) in enumerate(row, start=1):
            row_el.append(make_cell(cell_ref(c_idx, r_idx), val, style))
    dim = ET.Element(q("dimension"), {"ref": f"A1:{cell_ref(max(len(r) for r in rows), len(rows))}"})
    root.insert(0, dim)
    return root


def add_sheet(base, name, rows, styles, widths=None, freeze_ref=None):
    worksheets = base / "xl/worksheets"
    existing = sorted(worksheets.glob("sheet*.xml"))
    next_num = max(int(re.search(r"sheet(\d+)\.xml", p.name).group(1)) for p in existing) + 1
    sheet_file = worksheets / f"sheet{next_num}.xml"
    ET.ElementTree(build_sheet(rows, widths, freeze_ref)).write(sheet_file, encoding="utf-8", xml_declaration=True)

    wb_path = base / "xl/workbook.xml"
    wb = ET.parse(wb_path).getroot()
    sheets = wb.find(q("sheets"))
    rels_path = base / "xl/_rels/workbook.xml.rels"
    rels = ET.parse(rels_path).getroot()
    rid_nums = [int(re.search(r"\d+", r.get("Id")).group(0)) for r in rels if re.search(r"\d+", r.get("Id"))]
    rid = f"rId{max(rid_nums) + 1}"
    sheet_id = max(int(s.get("sheetId")) for s in sheets) + 1
    ET.SubElement(sheets, q("sheet"), {"name": name, "sheetId": str(sheet_id), f"{{{REL_NS}}}id": rid})
    ET.SubElement(rels, q("Relationship", PKG_REL_NS), {
        "Id": rid,
        "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
        "Target": f"worksheets/sheet{next_num}.xml",
    })
    ET.ElementTree(wb).write(wb_path, encoding="utf-8", xml_declaration=True)
    ET.ElementTree(rels).write(rels_path, encoding="utf-8", xml_declaration=True)

    ct_path = base / "[Content_Types].xml"
    ct = ET.parse(ct_path).getroot()
    ET.SubElement(ct, q("Override", CONTENT_NS), {
        "PartName": f"/xl/worksheets/sheet{next_num}.xml",
        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
    })
    ET.ElementTree(ct).write(ct_path, encoding="utf-8", xml_declaration=True)


def zip_xlsx(folder, out_path):
    if out_path.exists():
        out_path.unlink()
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in folder.rglob("*"):
            if path.is_file():
                zf.write(path, path.relative_to(folder).as_posix())


def achievement_style(actual, target, styles, rate=None):
    if target <= 0:
        return None
    rate = rate if rate is not None else actual / target
    if rate >= 1.00:
        return styles["green"]
    if rate >= 0.50:
        return styles["blue"]
    return None


def item_rate(item):
    return item.get("actual_rate") if item.get("actual_rate") is not None else (item["actual"] / item["target"] if item["target"] > 0 else 0)


def projected_rate(item):
    return item.get("rate", 0)


def diff_target_for_row(row, item):
    threshold = 0.80 if row["store"] == "北一二B整體" else 0.50
    return math.ceil(item["target"] * threshold)


def item_diff(row, item):
    return item["actual"] - diff_target_for_row(row, item)


def achievement_fill(actual, target, default="#ffffff"):
    if target <= 0:
        return default
    rate = actual / target
    if rate >= 1.00:
        return "#00b050"
    if rate >= 0.50:
        return "#43a5f5"
    return default


def progress_status(item):
    rate = item_rate(item)
    if item["target"] <= 0:
        return "無目標"
    if rate >= 1.00:
        return "超標"
    if rate >= 0.50:
        return "追蹤"
    return "落後"


def manager_award_tier(rate):
    if rate < 0.50:
        return None
    if rate < 0.80:
        return 0
    if rate < 1.00:
        return 1
    if rate < 1.20:
        return 2
    if rate < 2.00:
        return 3
    return 4


def supervisor_award_tier(rate):
    if rate < 0.80:
        return None
    if rate < 0.90:
        return 0
    if rate < 1.00:
        return 1
    if rate < 1.20:
        return 2
    if rate < 2.00:
        return 3
    return 4


def award_for_item(item, rules, tier_func):
    tier = tier_func(item_rate(item))
    if tier is None:
        return 0
    values = rules.get(item["name"])
    if not values or tier >= len(values):
        return 0
    return values[tier]


def manager_award_for_item(item):
    return award_for_item(item, MANAGER_AWARD_RULES, manager_award_tier)


def supervisor_award_for_item(item):
    return award_for_item(item, SUPERVISOR_AWARD_RULES, supervisor_award_tier)


def award_totals(row):
    return {
        "manager": sum(manager_award_for_item(item) for item in row["items"]),
        "supervisor": sum(supervisor_award_for_item(item) for item in row["items"]),
    }


MANAGER_TIER_LABELS = ["店50~79%", "店80~99%", "店100~119%", "店120~199%", "店>=200%"]
SUPERVISOR_TIER_LABELS = ["督80~89%", "督90~99%", "督100~119%", "督120~199%", "督>=200%"]


def award_rule_values(rules, group_name):
    values = rules.get(group_name, [])
    return [f"${value:,.0f}" for value in values]


def group_chunks(groups, size=5):
    return [groups[idx:idx + size] for idx in range(0, len(groups), size)]


def risk_threshold_for_store(store):
    return 0.80 if store == "北一二B整體" else 0.50


def display_item_name(name):
    if name in MODEL_DISPLAY_NAMES:
        return MODEL_DISPLAY_NAMES[name]
    display = {
        "Pixel10/10 Pro/10 Pro XL/10a": "Google Pixel 10／Pixel 10 Pro／Pixel 10 Pro XL／Pixel 10a",
        "razrfold": "moto razr fold",
        "S26Ultra/ZFold8/ZFold8Ultra": "Samsung Galaxy S26 Ultra／Z Fold8／Z Fold8 Ultra",
        "R11": "SHARP AQUOS R11",
        "X300/V70FE": "vivo X300／vivo V70 FE",
        "Pixel11Pro/11ProXL/11ProFold": "Google Pixel 11 Pro／Pixel 11 Pro XL／Pixel 11 Pro Fold",
        "S26/S26+/ZFlip8": "Samsung Galaxy S26／S26+／Z Flip8",
        "Pixel11": "Google Pixel 11",
        "Reno16F": "OPPO Reno16 F",
        "A57": "Samsung Galaxy A57",
        "A6x": "OPPO A6x 6G／128G",
        "A27/A17": "Samsung Galaxy A27／Samsung Galaxy A17",
        "Y21": "vivo Y21",
    }
    return display.get(name, name)


def summarize_store_progress(groups, data):
    rows = []
    for row in data:
        items = []
        risk_threshold = risk_threshold_for_store(row["store"])
        for item in row["items"]:
            rate = item_rate(item)
            gap_target = diff_target_for_row(row, item)
            items.append({
                "name": item["name"],
                "display_name": display_item_name(item["name"]),
                "actual": item["actual"],
                "target": item["target"],
                "rate": rate,
                "gap": max(gap_target - item["actual"], 0),
                "status": progress_status(item),
            })
        risks = [item for item in items if item["target"] > 0 and item["rate"] < risk_threshold]
        rows.append({
            "store": row["store"],
            "risk_threshold": risk_threshold,
            "risk_label": f"{int(risk_threshold * 100)}%",
            "items": items,
            "green_count": sum(1 for item in items if item["status"] == "超標"),
            "blue_count": sum(1 for item in items if item["status"] == "追蹤"),
            "lag_count": sum(1 for item in items if item["status"] == "落後"),
            "risk_count": len(risks),
            "strengths": sorted([item for item in items if item["status"] == "超標"], key=lambda item: item["rate"], reverse=True),
            "risks": sorted(risks, key=lambda item: (item["gap"], -item["rate"]), reverse=True),
        })
    return rows


def write_progress_markdown(progress, out_path, attachments):
    overall = progress[0] if progress else None
    stores = progress[1:]
    lag_sorted = sorted(stores, key=lambda row: (row["risk_count"], row["risks"][0]["gap"] if row["risks"] else 0, -row["green_count"]), reverse=True)
    strong_sorted = sorted(stores, key=lambda row: (row["green_count"], -row["lag_count"]), reverse=True)
    lines = [
        "# 北一二B 台獎機款進度點名",
        "",
        "## 整體狀況",
    ]
    if overall:
        lines.append(f"- 北一二B整體：超標 {overall['green_count']} 項、未達 80% {overall['risk_count']} 項。")
        if overall["risks"]:
            risks = "、".join(f"{item['display_name']} {item['rate']:.0%}（尚缺 {item['gap']:.0f} 台）" for item in overall["risks"][:5])
            lines.append(f"- 整體未達 80% 優先補量：{risks}")
    lines += [
        "",
        "## 店點點名",
    ]
    for row in lag_sorted:
        risk_text = "無明顯落後項"
        if row["risks"]:
            risk_text = "、".join(f"{item['display_name']} {item['rate']:.0%}（尚缺 {item['gap']:.0f} 台）" for item in row["risks"][:3])
        strength_text = "暫無超標項"
        if row["strengths"]:
            strength_text = "、".join(f"{item['display_name']} {item['rate']:.0%}" for item in row["strengths"][:3])
        lines.append(f"- {row['store']}：超標 {row['green_count']} 項，未達 50% {row['risk_count']} 項；需點名 {risk_text}；可複製強項 {strength_text}。")
    lines += [
        "",
        "## 今日優先順序",
    ]
    for row in lag_sorted[:3]:
        if row["risks"]:
            first = row["risks"][0]
            lines.append(f"- {row['store']} 先補 {first['display_name']}：目前 {first['actual']:.0f}/{first['target']:.0f}，達成 {first['rate']:.0%}。")
    lines += [
        "",
        "## 附件",
    ]
    for label, path in attachments.items():
        lines.append(f"- {label}：{path}")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def store_rows_for_xlsx(groups, data, styles):
    rows = [[("台獎機款", styles["header"])]]
    for chunk_idx, chunk in enumerate(group_chunks(groups), start=1):
        if chunk_idx > 1:
            rows.append([("", None)])
        header1 = [("店點", styles["header"]), ("實際獎金", styles["header"]), ("推估獎金", styles["header"]), ("排名", styles["header"]), ("領獎", styles["header"])]
        manager_tiers = [("", styles["subheader"])] * 5
        manager_values = [("", styles["subheader"])] * 5
        supervisor_tiers = [("", styles["subheader"])] * 5
        supervisor_values = [("", styles["subheader"])] * 5
        header2 = [("", styles["subheader"])] * 5
        for group in chunk:
            header1 += [(display_item_name(group["name"]), styles["header"]), ("", styles["header"]), ("", styles["header"]), ("", styles["header"])]
            manager_tiers += [(label, styles["subheader"]) for label in MANAGER_TIER_LABELS[:4]]
            manager_values += [(value, styles["subheader"]) for value in award_rule_values(MANAGER_AWARD_RULES, group["name"])[:4]]
            supervisor_tiers += [(label, styles["subheader"]) for label in SUPERVISOR_TIER_LABELS[:4]]
            supervisor_values += [(value, styles["subheader"]) for value in award_rule_values(SUPERVISOR_AWARD_RULES, group["name"])[:4]]
            header2 += [("實際數", styles["subheader"]), ("目標數", styles["subheader"]), ("差異", styles["subheader"]), ("預估達成", styles["subheader"])]
        rows += [header1, manager_tiers, manager_values, supervisor_tiers, supervisor_values, header2]
        chunk_names = {group["name"] for group in chunk}
        for row in data:
            not_award_style = styles["pink"] if row["award"] not in {"", None, "Y"} else None
            line = [
                (row["store"], not_award_style),
                (row["actual_total"], None),
                (row["projected"], None),
                (row["rank"], not_award_style),
                (row["award"], not_award_style),
            ]
            for item in [item for item in row["items"] if item["name"] in chunk_names]:
                actual, target, rate = item["actual"], item["target"], projected_rate(item)
                diff = item_diff(row, item)
                style = achievement_style(actual, target, styles, item_rate(item))
                line += [(actual, style), (target, style), (diff, styles["pink"] if diff < 0 else None), (round(rate, 4), None)]
            rows.append(line)
    return rows


def person_rows_for_xlsx(people, styles):
    headers = ["店點", "職稱", "員編", "姓名", "實際獎金", "推估獎金", "排名", "是否領獎", "銷售件數"]
    rows = [[("個人台獎", styles["header"])]]
    rows.append([(h, styles["header"]) for h in headers])
    for p in people:
        rows.append([
            (p["store"], None), (p["title"], None), (p["emp_id"], None), (p["name"], None),
            (p["actual_total"], styles["blue"] if p["actual_total"] >= 10000 else (styles["green"] if p["actual_total"] > 0 else None)),
            (p["projected"], None), (p["rank"], None), (p["award"], None), (p["units"], None)
        ])
    return rows


def html_escape(v):
    if isinstance(v, float):
        return f"{v:,.0f}" if v.is_integer() else f"{v:.1%}" if 0 < v < 4 else f"{v:,.1f}"
    return html.escape(str(v))


def write_store_html(groups, data, store_awards, out_path):
    tables = []
    for chunk in group_chunks(groups):
        rows = []
        chunk_names = {group["name"] for group in chunk}
        for row, award in zip(data, store_awards):
            not_award = award["award"] not in {"", None, "Y"}
            cells = f"<td class='{'award-no' if not_award else 'store'}'>{html_escape(row['store'])}</td>"
            award_flag_cls = "award-no" if not_award else ("award-yes" if award["award"] == "Y" else "")
            cells += (
                f"<td>{html_escape(display_amount(award['actual_total']))}</td>"
                f"<td>{html_escape(display_amount(award['projected']))}</td>"
                f"<td class='{'award-no' if not_award else ''}'>{award['rank']}</td>"
                f"<td class='{award_flag_cls}'>{award['award']}</td>"
            )
            for item in [item for item in row["items"] if item["name"] in chunk_names]:
                cls = ""
                if item["target"] > 0 and item_rate(item) >= 1.00:
                    cls = "green"
                elif item["target"] > 0 and item_rate(item) >= 0.50:
                    cls = "blue"
                diff = item_diff(row, item)
                cells += (
                    f"<td class='{cls}'>{item['actual']:.0f}</td><td class='{cls}'>{item['target']:.0f}</td>"
                    f"<td class='{'pink' if diff < 0 else ''}'>{diff:,.0f}</td><td>{projected_rate(item):.0%}</td>"
                )
            rows.append(f"<tr>{cells}</tr>")
        summary_cells = "<th rowspan='5'>實際獎金</th><th rowspan='5'>推估獎金</th><th rowspan='5'>排名</th><th rowspan='5'>是否領獎</th>"
        header1 = "<th rowspan='6'>店點</th><th colspan='4'>店點獎金摘要</th>" + "".join(f"<th colspan='4'>{html_escape(display_item_name(g['name']))}</th>" for g in chunk)
        manager_tiers = summary_cells + "".join("".join(f"<th>{label}</th>" for label in MANAGER_TIER_LABELS[:4]) for _ in chunk)
        manager_values = "".join("".join(f"<th>{value}</th>" for value in award_rule_values(MANAGER_AWARD_RULES, g["name"])[:4]) for g in chunk)
        supervisor_tiers = "".join("".join(f"<th>{label}</th>" for label in SUPERVISOR_TIER_LABELS[:4]) for _ in chunk)
        supervisor_values = "".join("".join(f"<th>{value}</th>" for value in award_rule_values(SUPERVISOR_AWARD_RULES, g["name"])[:4]) for g in chunk)
        header2 = "".join("<th>實際</th><th>目標</th><th>差異</th><th>達成</th>" for _ in chunk)
        tables.append(f"<table><thead><tr>{header1}</tr><tr>{manager_tiers}</tr><tr>{manager_values}</tr><tr>{supervisor_tiers}</tr><tr>{supervisor_values}</tr><tr>{header2}</tr></thead><tbody>{''.join(rows)}</tbody></table>")
    out_path.write_text(f"""<!doctype html><meta charset='utf-8'><style>
body{{margin:0;background:white;font-family:-apple-system,BlinkMacSystemFont,'PingFang TC','Microsoft JhengHei',Arial,sans-serif;}}
.wrap{{padding:14px;}}
h1{{font-size:26px;margin:0 0 12px;font-weight:900;}}
table{{border-collapse:collapse;font-size:14px;font-weight:800;margin-bottom:14px;}}
th,td{{border:1px solid #333;padding:6px 9px;text-align:center;white-space:nowrap;}}
th{{background:#ffd966;font-weight:900;}}
.store{{background:#c7f3f0;font-weight:900;text-align:left;}}
.blue{{background:#43a5f5;color:#001f3f;font-weight:900;}}
.green{{background:#00b050;color:#001b00;font-weight:900;}}
.pink{{background:#ffc7ce;color:#9c0006;font-weight:900;}}
.award-yes{{background:#e2f0d9;font-weight:900;}}
.award-no{{background:#ffc7ce;color:#9c0006;font-weight:900;}}
</style><div class='wrap'><h1>台獎機款</h1>{''.join(tables)}</div>""", encoding="utf-8")


def write_person_html(people, out_path):
    rows = []
    for p in people:
        cls = "pink" if p["award"] != "Y" else ("blue" if p["actual_total"] >= 10000 else "green")
        award_cls = "award-no" if p["award"] != "Y" else "award-yes"
        rows.append("<tr>" + "".join([
            f"<td class='store'>{html_escape(p['store'])}</td>",
            f"<td>{html_escape(p['name'])}</td>",
            f"<td class='{cls}'>{html_escape(display_amount(p['actual_total']))}</td>",
            f"<td>{html_escape(display_amount(p['projected']))}</td>",
            f"<td>{p['rank']}</td>",
            f"<td class='{award_cls}'>{html_escape(p['award'])}</td>",
        ]) + "</tr>")
    out_path.write_text(f"""<!doctype html><meta charset='utf-8'><style>
body{{margin:0;background:white;font-family:-apple-system,BlinkMacSystemFont,'PingFang TC','Microsoft JhengHei',Arial,sans-serif;}}
.wrap{{padding:14px;}}
h1{{font-size:26px;margin:0 0 12px;font-weight:900;}}
table{{border-collapse:collapse;font-size:20px;font-weight:800;}}
th,td{{border:1px solid #222;padding:7px 14px;text-align:center;white-space:nowrap;}}
th{{background:#00a6d6;color:white;font-weight:900;}}
.store{{background:#c7f3f0;font-weight:900;}}
.blue{{background:#8db4e2;font-weight:900;}}
.green{{background:#a9d18e;font-weight:900;}}
.pink{{background:#ffc7ce;color:#9c0006;font-weight:900;}}
.award-yes{{background:#e2f0d9;font-weight:900;}}
.award-no{{background:#ffc7ce;color:#9c0006;font-weight:900;}}
</style><div class='wrap'><h1>個人台獎</h1><table><thead><tr><th>店點</th><th>姓名</th><th>實際獎金</th><th>推估獎金</th><th>排名</th><th>是否領獎</th></tr></thead><tbody>{''.join(rows)}</tbody></table></div>""", encoding="utf-8")


def load_font(size, bold=False):
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold else "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/Microsoft JhengHei.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def text_size(draw, text, font):
    left, top, right, bottom = draw.textbbox((0, 0), str(text), font=font)
    return right - left, bottom - top


def draw_cell(draw, box, text, font, fill, outline, align="center", text_fill="#111111"):
    x1, y1, x2, y2 = box
    draw.rectangle(box, fill=fill, outline=outline, width=1)
    text = str(text)
    tw, th = text_size(draw, text, font)
    if align == "left":
        tx = x1 + 8
    elif align == "right":
        tx = x2 - tw - 8
    else:
        tx = x1 + (x2 - x1 - tw) / 2
    ty = y1 + (y2 - y1 - th) / 2 - 1
    draw.text((tx, ty), text, font=font, fill=text_fill)


def render_store_png(groups, data, store_awards, out_path):
    title_font = load_font(38, bold=True)
    header_font = load_font(18, bold=True)
    body_font = load_font(17, bold=True)
    small_font = load_font(16, bold=True)
    award_font = load_font(13, bold=True)

    store_width = 210
    summary_widths = [110, 110, 64, 72]
    item_widths = [76, 76, 84, 84]
    row_h = 40
    top_h = 70
    header1_h = 44
    award_h = 28
    header2_h = 38
    header_total_h = header1_h + award_h * 4 + header2_h
    block_gap = 24
    chunks = group_chunks(groups)
    block_widths = [store_width + sum(summary_widths) + sum(item_widths) * len(chunk) for chunk in chunks]
    block_h = header_total_h + row_h * len(data)
    width = max(block_widths) + 20
    height = top_h + block_h * len(chunks) + block_gap * (len(chunks) - 1) + 20
    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)

    draw.text((10, 12), "台獎機款", font=title_font, fill="#111111")

    y0 = top_h
    for chunk in chunks:
        chunk_names = {group["name"] for group in chunk}
        x = 10
        y = y0
        draw_cell(draw, (x, y, x + store_width, y + header_total_h), "店點", header_font, "#ffc000", "#333333")
        x += store_width
        summary_w = sum(summary_widths)
        draw_cell(draw, (x, y, x + summary_w, y + header1_h), "店點獎金摘要", small_font, "#00a6d6", "#333333", text_fill="#ffffff")
        summary_labels = ["實際獎金", "推估獎金", "排名", "領獎"]
        for label, w in zip(summary_labels, summary_widths):
            draw_cell(draw, (x, y + header1_h, x + w, y + header_total_h), label, header_font, "#ffc000", "#333333")
            x += w
        for group in chunk:
            group_w = sum(item_widths)
            draw_cell(draw, (x, y, x + group_w, y + header1_h), display_item_name(group["name"]), small_font, "#ffff99", "#333333")
            sub_y = y + header1_h
            for j, label in enumerate(MANAGER_TIER_LABELS[:4]):
                w = item_widths[j]
                draw_cell(draw, (x, sub_y, x + w, sub_y + award_h), label.replace("店", ""), award_font, "#ffffff", "#333333")
                x += w
            sub_y += award_h
            x -= group_w
            for j, value_text in enumerate(award_rule_values(MANAGER_AWARD_RULES, group["name"])[:4]):
                w = item_widths[j]
                draw_cell(draw, (x, sub_y, x + w, sub_y + award_h), value_text, award_font, "#ffffff", "#333333")
                x += w
            sub_y += award_h
            x -= group_w
            for j, label in enumerate(SUPERVISOR_TIER_LABELS[:4]):
                w = item_widths[j]
                draw_cell(draw, (x, sub_y, x + w, sub_y + award_h), label.replace("督", ""), award_font, "#ffffff", "#333333")
                x += w
            sub_y += award_h
            x -= group_w
            for j, value_text in enumerate(award_rule_values(SUPERVISOR_AWARD_RULES, group["name"])[:4]):
                w = item_widths[j]
                draw_cell(draw, (x, sub_y, x + w, sub_y + award_h), value_text, award_font, "#ffffff", "#333333")
                x += w
            sub_y += award_h
            x -= group_w
            labels = ["實際數", "目標數", "差異", "預估達成"]
            for j in range(4):
                w = item_widths[j]
                draw_cell(draw, (x, sub_y, x + w, sub_y + header2_h), labels[j], header_font, "#d9eaf7", "#333333")
                x += w

        y = y0 + header_total_h
        for row, award in zip(data, store_awards):
            x = 10
            store_fill = "#c7f3f0"
            if row["store"] == "北一二B整體":
                store_fill = "#ffc000"
            if award["award"] not in {"", None, "Y"}:
                store_fill = "#ffc7ce"
            draw_cell(draw, (x, y, x + store_width, y + row_h), row["store"], body_font, store_fill, "#333333", align="left")
            x += store_width
            award_text_fill = "#9c0006" if award["award"] and award["award"] != "Y" else "#111111"
            draw_cell(draw, (x, y, x + summary_widths[0], y + row_h), display_amount(award["actual_total"]), body_font, "#ffffff", "#333333", align="right")
            x += summary_widths[0]
            draw_cell(draw, (x, y, x + summary_widths[1], y + row_h), display_amount(award["projected"]), body_font, "#ffffff", "#333333", align="right")
            x += summary_widths[1]
            draw_cell(draw, (x, y, x + summary_widths[2], y + row_h), award["rank"] or "—", body_font, "#ffc7ce" if award["award"] not in {"", None, "Y"} else "#ffffff", "#333333", text_fill=award_text_fill)
            x += summary_widths[2]
            draw_cell(draw, (x, y, x + summary_widths[3], y + row_h), award["award"] or "—", body_font, "#ffc7ce" if award["award"] not in {"", None, "Y"} else ("#e2f0d9" if award["award"] == "Y" else "#ffffff"), "#333333", text_fill=award_text_fill)
            x += summary_widths[3]
            for item in [item for item in row["items"] if item["name"] in chunk_names]:
                rate = projected_rate(item)
                actual_fill = achievement_fill(item["actual"], item["target"])
                diff = item_diff(row, item)
                diff_text = f"{diff:,.0f}" if diff >= 0 else f"({abs(diff):,.0f})"
                draw_cell(draw, (x, y, x + item_widths[0], y + row_h), f"{item['actual']:.0f}", body_font, actual_fill, "#333333")
                x += item_widths[0]
                draw_cell(draw, (x, y, x + item_widths[1], y + row_h), f"{item['target']:.0f}", body_font, actual_fill, "#333333")
                x += item_widths[1]
                draw_cell(draw, (x, y, x + item_widths[2], y + row_h), diff_text, body_font, "#ffc7ce" if diff < 0 else "#ffffff", "#333333", text_fill="#9c0006" if diff < 0 else "#111111")
                x += item_widths[2]
                draw_cell(draw, (x, y, x + item_widths[3], y + row_h), f"{rate:.0%}", body_font, "#fff2cc", "#333333", text_fill="#9c0006" if rate < 0.5 else "#111111")
                x += item_widths[3]
            y += row_h
        y0 += block_h + block_gap

    image.save(out_path)


def render_person_png(people, out_path):
    title_font = load_font(34, bold=True)
    header_font = load_font(18, bold=True)
    body_font = load_font(17, bold=True)
    columns = [
        ("店點", 250),
        ("姓名", 100),
        ("實際獎金", 120),
        ("推估獎金", 140),
        ("排名", 80),
        ("是否領獎", 104),
    ]
    row_h = 40
    top_h = 66
    header_h = 42
    width = sum(w for _, w in columns) + 20
    height = top_h + header_h + row_h * len(people) + 20
    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)
    draw.text((10, 10), "個人台獎", font=title_font, fill="#111111")
    x = 10
    y = top_h
    for label, w in columns:
        draw_cell(draw, (x, y, x + w, y + header_h), label, header_font, "#00a6d6", "#222222", text_fill="#ffffff")
        x += w
    y += header_h
    for p in people:
        x = 10
        values = [
            p["store"], p["name"], display_amount(p["actual_total"]),
            display_amount(p["projected"]), p["rank"] or "—", p["award"] or "—"
        ]
        fills = [
            "#c7f3f0",
            "#ffffff",
            "#ffc7ce" if p["award"] != "Y" else ("#8db4e2" if p["actual_total"] >= 10000 else "#a9d18e"),
            "#ffffff",
            "#ffffff",
            "#ffc7ce" if p["award"] != "Y" else "#e2f0d9",
        ]
        text_fills = ["#111111", "#111111", "#9c0006" if p["award"] != "Y" else "#111111", "#111111", "#111111", "#9c0006" if p["award"] != "Y" else "#111111"]
        aligns = ["left", "center", "right", "right", "center", "center"]
        for (label, w), value, fill, text_fill, align in zip(columns, values, fills, text_fills, aligns):
            draw_cell(draw, (x, y, x + w, y + row_h), value, body_font, fill, "#222222", align=align, text_fill=text_fill)
            x += w
        y += row_h
    image.save(out_path)


def main():
    global STORE_SRC, PERSON_SRC, Y26_SRC, MANAGER_AWARD_RULES, SUPERVISOR_AWARD_RULES, Image, ImageDraw, ImageFont
    apply_award_config(load_active_award_config())
    kpi_identity = kpi_identity_for_batch(required_report_run_date())
    expected_data_date = kpi_identity["source_data_date"]
    store_source, store_origin = resolve_exact_or_fresh_source(
        EXACT_STORE_SOURCE, "01-08-03", "store", expected_data_date,
    )
    person_source, person_origin = resolve_exact_or_fresh_source(
        EXACT_PERSON_SOURCE, "01-08-04", "person", expected_data_date,
    )
    freshness = validate_fresh_award_sources(store_source, person_source, store_origin, person_origin)
    if os.environ.get("PHONE_AWARDS_VERIFY_SOURCE_ONLY") == "1":
        for kind, source in (("store", store_source), ("person", person_source)):
            evidence = source_date_provenance_from_xlsx(source, kind)
            freshness["source_identity"][kind]["source_date_range"] = evidence["source_date_range"]
            freshness["source_identity"][kind]["source_data_date"] = source_data_date_from_range(
                evidence["source_date_range"], required_report_run_date(),
            )
            freshness["source_identity"][kind]["date_provenance"] = evidence["date_provenance"]
        print(json.dumps({
            "status": freshness["status"],
            "run_id": freshness["run_id"],
            "report_run_date": freshness["report_run_date"],
            "data_cutoff_date": expected_data_date,
            "kpi_source_identity": kpi_identity,
            "source_identity": freshness["source_identity"],
        }, ensure_ascii=False))
        return

    source_batch = validate_same_source_batch(freshness, store_source, person_source)
    if os.environ.get("PHONE_AWARDS_VERIFY_BATCH_ONLY") == "1":
        print(json.dumps({"status": "ok", "source_batch": source_batch}, ensure_ascii=False))
        return

    # Keep the fail-closed source-only preflight usable in minimal runners.  PNG
    # rendering remains mandatory for a real build, but must not be imported
    # before we reject stale/missing inputs.
    from PIL import Image, ImageDraw, ImageFont

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    y26_source = resolve_y26_source()
    STORE_SRC = unpack_or_copy_workbook(store_source, OUT_DIR / "01-08-03_phone_awards_source_package")
    PERSON_SRC = unpack_or_copy_workbook(person_source, OUT_DIR / "01-08-04_phone_awards_source_package")
    Y26_SRC = unpack_or_copy_workbook(y26_source, OUT_DIR / "Y26_phone_awards_source_package") if y26_source.exists() else y26_source

    store_groups, store_data = build_store_awards(source="source")
    people = build_person_awards(source="source")
    if not Y26_SRC.exists():
        raise FileNotFoundError("Y26 source is required to load current manager and supervisor award rows")
    MANAGER_AWARD_RULES, SUPERVISOR_AWARD_RULES = load_central_award_rules(store_groups)
    store_awards = [
        {
            "store": row["store"],
            "actual_total": row["actual_total"],
            "projected": row["projected"],
            "rank": row["rank"],
            "award": row["award"],
            "manager_award": award_totals(row)["manager"],
            "supervisor_award": award_totals(row)["supervisor"],
        }
        for row in store_data
    ]

    y26_xlsx = None
    if Y26_SRC.exists():
        y26_work = OUT_DIR / "Y2606_phone_awards_package"
        copy_tree(Y26_SRC, y26_work)
        update_y26_workbook(y26_work, store_groups, store_data, people)
        y26_xlsx = OUT_DIR / "Y26重點台獎手機_台獎更新.xlsx"
        zip_xlsx(y26_work, y26_xlsx)

    store_work = OUT_DIR / "01-08-03_phone_awards_package"
    person_work = OUT_DIR / "01-08-04_phone_awards_package"
    copy_tree(STORE_SRC, store_work)
    copy_tree(PERSON_SRC, person_work)

    store_styles = ensure_styles(store_work)
    person_styles = ensure_styles(person_work)
    add_sheet(store_work, "台獎機款", store_rows_for_xlsx(store_groups, store_data, store_styles), store_styles, widths=[18, 13, 13, 8, 8] + [9, 9, 9, 9] * len(store_groups), freeze_ref="A8")
    add_sheet(person_work, "個人台獎", person_rows_for_xlsx(people, person_styles), person_styles, widths=[26, 14, 12, 12, 12, 14, 10, 12, 12], freeze_ref="A3")

    store_xlsx = OUT_DIR / "01-08-03_手機競賽日報_店點達成率_台獎機款更新.xlsx"
    person_xlsx = OUT_DIR / "01-08-04_手機競賽日報_個人達成率_個人台獎更新.xlsx"
    zip_xlsx(store_work, store_xlsx)
    zip_xlsx(person_work, person_xlsx)

    store_html = OUT_DIR / "台獎機款.html"
    person_html = OUT_DIR / "個人台獎.html"
    write_store_html(store_groups, store_data, store_awards, store_html)
    write_person_html(people, person_html)
    store_png = OUT_DIR / "台獎機款.png"
    person_png = OUT_DIR / "個人台獎.png"
    render_store_png(store_groups, store_data, store_awards, store_png)
    render_person_png(people, person_png)
    progress_md = OUT_DIR / "台獎機款_店點進度點名.md"
    progress = summarize_store_progress(store_groups, store_data)
    progress_attachments = {
        "Y26台獎更新報表": y26_xlsx,
        "台獎機款店點報表": store_xlsx,
        "個人台獎報表": person_xlsx,
        "台獎機款HTML": store_html,
        "個人台獎HTML": person_html,
        "台獎機款截圖": store_png,
        "個人台獎截圖": person_png,
    }
    write_progress_markdown(progress, progress_md, {k: v for k, v in progress_attachments.items() if v})

    summary = {
        "run_id": freshness["run_id"],
        "report_run_date": freshness["report_run_date"],
        "report_run_date": freshness["report_run_date"],
        "config_version": ACTIVE_AWARD_CONFIG.get("configVersion") if ACTIVE_AWARD_CONFIG else "legacy-hardcoded",
        "config_path": ACTIVE_AWARD_CONFIG.get("_path") if ACTIVE_AWARD_CONFIG else None,
        "config_hash": ACTIVE_AWARD_CONFIG.get("_hash") if ACTIVE_AWARD_CONFIG else None,
        "effective_month": ACTIVE_AWARD_CONFIG.get("effectiveMonth") if ACTIVE_AWARD_CONFIG else None,
        "store_xlsx": str(store_xlsx),
        "person_xlsx": str(person_xlsx),
        "store_html": str(store_html),
        "person_html": str(person_html),
        "store_png": str(store_png),
        "person_png": str(person_png),
        "progress_md": str(progress_md),
        "y26_xlsx": str(y26_xlsx) if y26_xlsx else None,
        "source_files": {
            "store": str(store_source),
            "person": str(person_source),
            "y26": str(y26_source),
        },
        "source_hashes": {
            "store": freshness["source_identity"]["store"]["sha256"],
            "person": freshness["source_identity"]["person"]["sha256"],
            "y26": file_sha256(y26_source) if y26_source.exists() else None,
        },
        "source_identity": freshness["source_identity"],
        "source_freshness": {
            "status": freshness["status"],
            "previous": freshness["previous"],
            "unchanged_exception": freshness["unchanged_exception"],
        },
        "source_batch": source_batch,
        "source_mode": "fresh 01-08-03/01-08-04 -> Y26 tabs -> screenshots" if y26_xlsx else "fresh 01-08-03/01-08-04 screenshots",
        "store_rows": len(store_data),
        "phone_items": len(store_groups),
        "person_award_y_rows": sum(1 for p in people if p["award"] == "Y"),
        "store_progress": progress,
        "store_awards": store_awards,
        "person_awards": people,
        "award_rules": {
            "manager": MANAGER_AWARD_RULES,
            "supervisor": SUPERVISOR_AWARD_RULES,
            "manager_tiers": ["50%-79%", "80%-99%", "100%-119%", "120%-199%", ">=200%"],
            "supervisor_tiers": ["80%-89%", "90%-99%", "100%-119%", "120%-199%", ">=200%"],
            "source": {
                "workbook": str(y26_source),
                "sheet": "台獎機款",
                "manager_amount_rows": [3, 19],
                "supervisor_amount_rows": [5, 21],
            },
        },
        "top_person": people[0] if people else None,
    }
    (OUT_DIR / "phone_awards_update_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
