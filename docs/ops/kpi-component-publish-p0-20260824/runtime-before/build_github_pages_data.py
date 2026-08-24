#!/usr/bin/env python3
"""Build compact website data from the OneDrive daily reports.

The raw OneDrive workbooks stay untouched. This script extracts only the
daily-momentum metrics needed by the GitHub Pages dashboard and the latest
validated 13-item phone-awards summary.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from calendar import monthrange
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from build_personal_kpi_report import apply_previous_rates, extract_rows


STORE_ORDER = [
    "通化", "酒泉", "台北三創", "萬大", "六張犁",
    "復興南", "永吉", "大稻埕", "杭州南",
]

METRICS = {
    "5G銷售數": ("five_g", "5G", "筆"),
    "TTL AQ上線點數": ("aq_points", "AQ", "點"),
    "TTL AQ上線數": ("aq_count", "AQ上線數", "筆"),
    "AQ V+D 999 (含)以上": ("a999", "A999", "筆"),
    "AQ V+D 1399 (含)以上": ("a1399", "A1399", "筆"),
    "RT上線點數": ("rt_points", "RT", "點"),
    "RT上線數": ("rt_count", "RT上線數", "筆"),
    "RT V+D 999 (含)以上": ("r999", "R999", "筆"),
    "RT V+D 1399 (含)以上": ("r1399", "R1399", "筆"),
    "好速案銷售點數": ("haosu", "好速", "點"),
}

SOURCE_RE = re.compile(r"^(\d{4})\.xlsx$")
REQUIRED_AWARD_SOURCE_BASENAMES = {
    "store": "01-08-03-(密)直營_手機競賽日報_店點達成率、排名及獎金.xlsx",
    "person": "01-08-04-(密)直營_手機競賽日報_個人達成率、排名及獎金.xlsx",
}

AWARD_STEPS = [
    {"threshold": 0.50, "manager_tier": 0, "supervisor_tier": None, "label": "解鎖店長50%"},
    {"threshold": 0.80, "manager_tier": 1, "supervisor_tier": 0, "label": "衝督導80%"},
    {"threshold": 0.90, "manager_tier": 1, "supervisor_tier": 1, "label": "升督導90%"},
    {"threshold": 1.00, "manager_tier": 2, "supervisor_tier": 2, "label": "衝100%"},
    {"threshold": 1.20, "manager_tier": 3, "supervisor_tier": 3, "label": "衝120%"},
    {"threshold": 2.00, "manager_tier": 4, "supervisor_tier": 4, "label": "衝200%"},
]

KPI_BATTLE_CORE_METRICS = {
    "a999": ("A999", ["AQ V+D 999（含）以上", "AQ V+D 999 (含)以上"]),
    "a1399": ("A1399", ["AQ V+D 1399（含）以上", "AQ V+D 1399 (含)以上"]),
    "haosu": ("好速", ["好速案銷售點數"]),
    "r1399": ("R1399", ["RT V+D 1399（含）以上", "RT V+D 1399 (含)以上"]),
}


def iso_from_value(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip().replace("/", "-")
        match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
        if match:
            return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return None


def numeric(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = str(value).replace(",", "").strip()
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def normalize_store(raw: str) -> str:
    text = raw.strip()
    if "北一二B" in text or "北一二Ｂ" in text:
        return "北一二B整體"
    if "三創" in text:
        return "台北三創"
    text = text.replace("台灣大哥大數位生活", "")
    if text.startswith("台北"):
        text = text[2:]
    return text


def source_files(source_dir: Path) -> list[tuple[tuple[int, int], Path]]:
    candidates: list[tuple[tuple[int, int], Path]] = []
    for path in source_dir.glob("*.xlsx"):
        match = SOURCE_RE.match(path.name)
        if not match:
            continue
        mmdd = match.group(1)
        candidates.append(((int(mmdd[:2]), int(mmdd[2:])), path))
    return sorted(candidates)


def canonical_input_source_file(report: dict[str, Any]) -> str:
    raw = str(report.get("source_path") or report.get("source_file") or "").strip()
    source_file = Path(raw.replace("\\", "/")).name
    if not source_file or source_file in {".", ".."} or Path(source_file).suffix.lower() != ".xlsx":
        raise RuntimeError("today_report_data.json 缺少 canonical source_path/source_file")
    return source_file


def extract_momentum(source_dir: Path) -> dict[str, Any]:
    data: dict[str, dict[str, dict[str, int | float]]] = {}
    latest_source_by_date: dict[str, str] = {}
    parsed_files: list[str] = []

    for (_, source_path) in source_files(source_dir):
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            if "上線數KPI_每日上線" not in workbook.sheetnames:
                continue
            rows = list(workbook["上線數KPI_每日上線"].iter_rows(values_only=True))
        except Exception:
            continue
        if len(rows) < 9:
            continue

        date_columns: list[tuple[int, str]] = []
        for col_idx, value in enumerate(rows[7], start=0):
            iso = iso_from_value(value)
            if iso:
                date_columns.append((col_idx, iso))
        if not date_columns:
            continue

        block_starts: list[tuple[int, str]] = []
        for row_idx, row in enumerate(rows):
            raw = row[0] if row else None
            if not isinstance(raw, str):
                continue
            text = raw.strip()
            if "北一二B" in text or "北一二Ｂ" in text or text.startswith("台北") or text.startswith("台灣大哥大"):
                block_starts.append((row_idx, normalize_store(text)))
        if not block_starts:
            continue

        parsed_files.append(source_path.name)
        for block_idx, (start_row, store) in enumerate(block_starts):
            end_row = block_starts[block_idx + 1][0] if block_idx + 1 < len(block_starts) else len(rows)
            store_dates = data.setdefault(store, {})
            for row in rows[start_row + 1:end_row]:
                if not row or not isinstance(row[0], str):
                    continue
                metric = METRICS.get(row[0].strip())
                if not metric:
                    continue
                key = metric[0]
                for col_idx, iso in date_columns:
                    value = numeric(row[col_idx] if col_idx < len(row) else None)
                    if value is None:
                        continue
                    store_dates.setdefault(iso, {})[key] = value
                    latest_source_by_date[iso] = source_path.name

    stores = ["北一二B整體"] + [s for s in STORE_ORDER if s in data]
    for store in data:
        if store not in stores:
            stores.append(store)
    metric_meta = {
        key: {"label": label, "unit": unit}
        for key, label, unit in METRICS.values()
    }
    return {
        "version": 1,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_sheet": "上線數KPI_每日上線",
        "stores": stores,
        "metrics": metric_meta,
        "data": data,
        "source_files": parsed_files,
        "latest_source_by_date": latest_source_by_date,
    }


def validated_award_source_files(summary: dict[str, Any], report_run_date: str) -> dict[str, dict[str, str]]:
    """Carry immutable per-report award evidence; a KPI filename is never an award source."""
    identity = summary.get("source_identity")
    if not isinstance(identity, dict):
        raise RuntimeError("台獎 source_identity 缺少，禁止建立 snapshot")
    run_id = str(summary.get("run_id") or "")
    if not run_id or str(summary.get("report_run_date") or "") != report_run_date:
        raise RuntimeError("台獎 run_id/report_run_date 不完整，禁止建立 snapshot")
    freshness = summary.get("source_freshness") or {}
    if freshness.get("status") not in {"fresh", "unchanged-source-verified-exception"}:
        raise RuntimeError("台獎來源 freshness 未驗證，禁止建立 snapshot")

    source_files: dict[str, dict[str, str]] = {}
    for kind, expected_basename in REQUIRED_AWARD_SOURCE_BASENAMES.items():
        entry = identity.get(kind)
        if not isinstance(entry, dict):
            raise RuntimeError(f"台獎 {kind} source identity 缺少")
        basename = str(entry.get("canonical_basename") or "")
        sha256 = str(entry.get("sha256") or "").lower()
        mtime = str(entry.get("mtime") or "")
        if basename != expected_basename:
            raise RuntimeError(f"台獎 {kind} canonical basename 不符：{basename}")
        if not re.fullmatch(r"[a-f0-9]{64}", sha256):
            raise RuntimeError(f"台獎 {kind} SHA-256 缺少")
        if not mtime or str(entry.get("run_id") or "") != run_id:
            raise RuntimeError(f"台獎 {kind} mtime/run_id 缺少")
        source_files[kind] = {
            "basename": basename,
            "sha256": sha256,
            "mtime": mtime,
            "run_id": run_id,
        }
    return source_files


def extract_phone_awards(
    summary_path: Path,
    report_run_date: str,
    data_cutoff_date: str,
) -> dict[str, Any]:
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    source_files = validated_award_source_files(summary, report_run_date)
    progress = summary.get("store_progress", [])
    overall = next((row for row in progress if row.get("store") == "北一二B整體"), None)
    if overall is None and progress:
        overall = progress[0]
    overall = overall or {}

    item_rows: dict[str, dict[str, Any]] = {}
    for row in (overall.get("strengths", []) + overall.get("risks", [])):
        item_rows[row.get("name", row.get("display_name", ""))] = row

    return {
        "version": 1,
        "report_date": data_cutoff_date,
        "report_run_date": report_run_date,
        "data_as_of_date": data_cutoff_date,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_mode": summary.get("source_mode"),
        "source_files": source_files,
        "phone_items": summary.get("phone_items"),
        "store_rows": summary.get("store_rows"),
        "person_award_y_rows": summary.get("person_award_y_rows"),
        "items": list(item_rows.values()),
        "stores": progress,
    }


def award_amount(rules: dict[str, Any], item_name: str, tier: int | None) -> int | float:
    if tier is None:
        return 0
    values = rules.get(item_name, [])
    return numeric(values[tier]) or 0 if tier < len(values) else 0


def current_award_tier(rate: float, role: str) -> int | None:
    applicable = [step for step in AWARD_STEPS if step["threshold"] <= rate]
    if not applicable:
        return None
    return applicable[-1][f"{role}_tier"]


def phone_award_priority(item: dict[str, Any], rules: dict[str, Any]) -> dict[str, Any] | None:
    actual = numeric(item.get("actual"))
    target = numeric(item.get("target"))
    rate = numeric(item.get("rate"))
    name = str(item.get("name", ""))
    if actual is None or not target or rate is None or not name:
        return None
    next_step = next((step for step in AWARD_STEPS if rate < step["threshold"] - 1e-9), None)
    if not next_step:
        return None
    units_needed = max(math.ceil(target * next_step["threshold"] - actual - 1e-9), 0)
    if units_needed <= 0:
        return None
    manager_rules = rules.get("manager", {})
    supervisor_rules = rules.get("supervisor", {})
    manager_current = award_amount(manager_rules, name, current_award_tier(float(rate), "manager"))
    supervisor_current = award_amount(supervisor_rules, name, current_award_tier(float(rate), "supervisor"))
    manager_next = award_amount(manager_rules, name, next_step["manager_tier"])
    supervisor_next = award_amount(supervisor_rules, name, next_step["supervisor_tier"])
    manager_gain = max(manager_next - manager_current, 0)
    supervisor_gain = max(supervisor_next - supervisor_current, 0)
    incremental = manager_gain + supervisor_gain
    close_unlock = rate < 0.50 and units_needed <= 3
    return {
        **item,
        "units_needed": units_needed,
        "next_threshold": next_step["threshold"],
        "next_label": next_step["label"],
        "manager_current": manager_current,
        "supervisor_current": supervisor_current,
        "manager_gain": manager_gain,
        "supervisor_gain": supervisor_gain,
        "incremental_award": incremental,
        "award_per_unit": round(incremental / units_needed, 1),
        "close_unlock": close_unlock,
    }


def monthly_award_max(item_name: str, rules: dict[str, Any]) -> int | float:
    """Fixed monthly model priority: highest available manager + supervisor award."""
    manager = [numeric(value) or 0 for value in rules.get("manager", {}).get(item_name, [])]
    supervisor = [numeric(value) or 0 for value in rules.get("supervisor", {}).get(item_name, [])]
    return (max(manager) if manager else 0) + (max(supervisor) if supervisor else 0)


def award_at_threshold(rules: dict[str, Any], item_name: str, role: str, threshold: float) -> int | float:
    """Return one role's configured amount at an exact monthly award threshold."""
    step = next((item for item in AWARD_STEPS if abs(item["threshold"] - threshold) < 1e-9), None)
    return award_amount(rules.get(role, {}), item_name, step.get(f"{role}_tier") if step else None)


def enrich_phone_award_item(item: dict[str, Any], rules: dict[str, Any], threshold: float) -> dict[str, Any]:
    actual = numeric(item.get("actual"))
    target = numeric(item.get("target"))
    rate = numeric(item.get("rate"))
    name = str(item.get("name", ""))
    threshold_target = math.ceil(float(target) * threshold) if target else None
    difference = actual - threshold_target if actual is not None and threshold_target is not None else None
    next_award = phone_award_priority(item, rules)
    return {
        **item,
        "actual": actual,
        "target": target,
        "rate": rate,
        "threshold": threshold,
        "threshold_target": threshold_target,
        "difference": difference,
        "monthly_award_max": monthly_award_max(name, rules),
        "units_needed": (next_award or {}).get("units_needed"),
        "next_threshold": (next_award or {}).get("next_threshold"),
        "next_label": (next_award or {}).get("next_label"),
        "incremental_award": (next_award or {}).get("incremental_award", 0),
        # The selector presents the role-specific tiers only: store manager 50/100,
        # North12B supervisor 80/100. Do not combine the two award tracks here.
        "store_reward_50": award_at_threshold(rules, name, "manager", 0.50),
        "store_reward_100": award_at_threshold(rules, name, "manager", 1.00),
        "district_reward_80": award_at_threshold(rules, name, "supervisor", 0.80),
        "district_reward_100": award_at_threshold(rules, name, "supervisor", 1.00),
    }


def monthly_priority_key(item: dict[str, Any]) -> tuple[float, str]:
    return (-float(numeric(item.get("monthly_award_max")) or 0), str(item.get("display_name") or item.get("name") or ""))


def extract_phone_awards_battle(
    summary_path: Path,
    report_run_date: str,
    data_cutoff_date: str,
) -> dict[str, Any]:
    """Private action view: actual/projected awards and fixed monthly phone priorities."""
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    source_files = validated_award_source_files(summary, report_run_date)
    rules = summary.get("award_rules", {})
    awards_by_store = {
        normalize_store(str(row.get("store", ""))): row
        for row in summary.get("store_awards", [])
    }
    stores = []
    overall = None
    for row in summary.get("store_progress", []):
        store = normalize_store(str(row.get("store", "")))
        threshold = 0.80 if store == "北一二B整體" else 0.50
        items = sorted(
            [enrich_phone_award_item(item, rules, threshold) for item in row.get("items", [])],
            key=monthly_priority_key,
        )
        # User rule: only rates strictly over 100% are skipped; exactly 100% stays.
        priorities = [item for item in items if numeric(item.get("rate")) is not None and float(item["rate"]) <= 1.0 + 1e-9][:3]
        compact = {
            "store": store,
            "award": awards_by_store.get(store, {}),
            "items": items,
            "priorities": priorities,
        }
        if store == "北一二B整體":
            overall = compact
        else:
            stores.append(compact)
    stores.sort(key=lambda row: (
        -float(numeric(row.get("award", {}).get("actual_total")) or 0),
        float(numeric(row.get("award", {}).get("rank")) or 99999),
        row.get("store", ""),
    ))
    return {
        "version": 2,
        "visibility": "private-local-preview",
        "report_date": data_cutoff_date,
        "report_run_date": report_run_date,
        "data_as_of_date": data_cutoff_date,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_mode": summary.get("source_mode"),
        "source_files": source_files,
        "phone_items": summary.get("phone_items"),
        "store_rows": summary.get("store_rows"),
        "supervisor": (overall or {}).get("award", {}),
        "overall": overall or {},
        "stores": stores,
    }


def mask_name(value: Any) -> str:
    """Keep the local dashboard useful without exposing a full employee name."""
    name = str(value or "").strip()
    if not name:
        return "—"
    if "*" in name or "＊" in name:
        return name
    if len(name) == 1:
        return "＊"
    if len(name) == 2:
        return f"{name[0]}＊"
    return f"{name[0]}＊{name[-1]}"


def find_metric(source: dict[str, Any], candidates: list[str]) -> dict[str, Any]:
    for candidate in candidates:
        if candidate in source:
            return source[candidate]
    normalized = {re.sub(r"\s+", "", str(key)): value for key, value in source.items()}
    for candidate in candidates:
        hit = normalized.get(re.sub(r"\s+", "", candidate))
        if hit is not None:
            return hit
    return {"actual": None, "target": None, "achievement": None}


def source_as_of_date(value: Any, fallback: date) -> date:
    """Use the last date in the source range, never the dashboard viewing date."""
    matches = re.findall(r"(?:(\d{4})[/-])?(\d{1,2})[/-](\d{1,2})", str(value or ""))
    if not matches:
        return fallback
    year_text, month_text, day_text = matches[-1]
    try:
        return date(int(year_text or fallback.year), int(month_text), int(day_text))
    except ValueError:
        return fallback


def resolve_date_contract(
    report: dict[str, Any],
    report_run_date_value: str | None = None,
    data_cutoff_date_value: str | None = None,
) -> tuple[date, date]:
    """Separate mail/run date from the formal website data date, fail closed."""
    report_run_text = str(report.get("report_date_iso") or "")
    if not report_run_text:
        raise RuntimeError("today_report_data.json 缺少 report_date_iso")
    try:
        report_run_date = date.fromisoformat(report_run_text)
    except ValueError as exc:
        raise RuntimeError(f"today_report_data.json report_date_iso 無效：{report_run_text}") from exc
    if report_run_date_value:
        try:
            requested_run_date = date.fromisoformat(report_run_date_value)
        except ValueError as exc:
            raise RuntimeError(f"--report-run-date 無效：{report_run_date_value}") from exc
        if requested_run_date != report_run_date:
            raise RuntimeError(
                f"report run date mismatch: expected {requested_run_date}, got {report_run_date}"
            )

    source_range = report.get("source_date_range")
    if not re.findall(r"(?:(\d{4})[/-])?(\d{1,2})[/-](\d{1,2})", str(source_range or "")):
        raise RuntimeError(f"無法從 source_date_range 取得資料截止日：{source_range}")
    parsed_cutoff = source_as_of_date(source_range, report_run_date)
    if data_cutoff_date_value:
        try:
            requested_cutoff = date.fromisoformat(data_cutoff_date_value)
        except ValueError as exc:
            raise RuntimeError(f"--data-cutoff-date 無效：{data_cutoff_date_value}") from exc
        if requested_cutoff != parsed_cutoff:
            raise RuntimeError(
                f"data cutoff date mismatch: expected {requested_cutoff}, got {parsed_cutoff}"
            )
    if parsed_cutoff > report_run_date:
        raise RuntimeError(
            f"資料截止日 {parsed_cutoff} 不得晚於 report run date {report_run_date}"
        )
    return report_run_date, parsed_cutoff


def compact_metric(
    value: dict[str, Any],
    previous: dict[str, Any] | None = None,
    as_of_date: date | None = None,
) -> dict[str, int | float | None]:
    actual = numeric(value.get("actual"))
    target = numeric(value.get("target"))
    rate = numeric(value.get("achievement", value.get("rate")))
    previous_rate = numeric((previous or {}).get("achievement", (previous or {}).get("rate")))
    daily_target = None
    daily_gap = None
    if target is not None and as_of_date:
        daily_target = round(target * as_of_date.day / monthrange(as_of_date.year, as_of_date.month)[1], 1)
        daily_gap = round(actual - daily_target, 1) if actual is not None else None
    return {
        "actual": actual,
        "target": target,
        "daily_target": daily_target,
        "daily_gap": daily_gap,
        "rate": rate,
        "dod": round(rate - previous_rate, 4) if rate is not None and previous_rate is not None else None,
    }


def extract_kpi_battle(
    today_report_path: Path,
    phone_summary_path: Path | None = None,
    supplemental_summary_path: Path | None = None,
    report_run_date_value: str | None = None,
    data_cutoff_date_value: str | None = None,
) -> dict[str, Any]:
    """Build a private, latest-day supervisor view from the official report artifacts."""
    report = json.loads(today_report_path.read_text(encoding="utf-8"))
    report_run_date, as_of_date = resolve_date_contract(
        report,
        report_run_date_value,
        data_cutoff_date_value,
    )
    previous_date = as_of_date - timedelta(days=1)
    previous_report_path = today_report_path.parent / f"report_data_{previous_date.isoformat()}.json"
    previous_report = (
        json.loads(previous_report_path.read_text(encoding="utf-8"))
        if previous_report_path.exists() else None
    )
    source_path = Path(str(report.get("source_path", "")))
    if not source_path.exists():
        raise RuntimeError(f"個績來源不存在：{source_path}")

    # Actual insurance attachment rate is produced by the same report run but is
    # not part of today_report_data. Refuse mismatched summary metadata so an old
    # supplemental output can never be merged into the private snapshot.
    insurance_total_rate = None
    insurance_rates_by_store: dict[str, float] = {}
    if supplemental_summary_path and supplemental_summary_path.exists():
        supplemental = json.loads(supplemental_summary_path.read_text(encoding="utf-8"))
        if (
            str(supplemental.get("report_date_iso", "")) == str(report.get("report_date_iso", ""))
            and str(supplemental.get("source_file", "")) == source_path.name
        ):
            insurance_total_rate = numeric((supplemental.get("insurance") or {}).get("attach_rate"))
            detail_rows = ((supplemental.get("details") or {}).get("insurance") or {}).get("rows") or {}
            insurance_rates_by_store = {
                normalize_store(str(store)): numeric((values or {}).get("attach_rate"))
                for store, values in detail_rows.items()
                if numeric((values or {}).get("attach_rate")) is not None
            }

    previous_by_store: dict[str, dict[str, Any]] = {}
    for previous in [previous_report.get("aggregate") if previous_report else None, *((previous_report or {}).get("records", []))]:
        if not previous:
            continue
        for key in (previous.get("store"), previous.get("display_store")):
            if key:
                previous_by_store[normalize_store(str(key))] = previous

    def compact_store(row: dict[str, Any]) -> dict[str, Any]:
        previous = previous_by_store.get(normalize_store(str(row.get("display_store") or row.get("store", ""))))
        other_metrics = row.get("other_metrics", {})
        previous_other_metrics = (previous or {}).get("other_metrics", {})
        core = {
            key: {
                "label": label,
                **compact_metric(
                    find_metric(other_metrics, names),
                    find_metric(previous_other_metrics, names),
                    as_of_date,
                ),
            }
            for key, (label, names) in KPI_BATTLE_CORE_METRICS.items()
        }
        overall_kpi = numeric(row.get("overall_kpi"))
        previous_overall_kpi = numeric((previous or {}).get("overall_kpi"))
        company_rank = numeric(row.get("company_rank"))
        previous_company_rank = numeric((previous or {}).get("company_rank"))
        addon_score = numeric(row.get("addon_score"))
        previous_addon_score = numeric((previous or {}).get("addon_score"))
        normalized_store = normalize_store(str(row.get("display_store") or row.get("store", "")))
        return {
            "store": row.get("display_store") or normalized_store,
            "company_rank": company_rank,
            "company_rank_dod": int(previous_company_rank - company_rank) if company_rank is not None and previous_company_rank is not None else None,
            "overall_kpi": overall_kpi,
            "overall_kpi_dod": round(overall_kpi - previous_overall_kpi, 4) if overall_kpi is not None and previous_overall_kpi is not None else None,
            "addon_score": addon_score,
            "addon_score_dod": round(addon_score - previous_addon_score, 2) if addon_score is not None and previous_addon_score is not None else None,
            "insurance_attach_rate": insurance_total_rate if normalized_store == "北一二B整體" else insurance_rates_by_store.get(normalized_store),
            "core": core,
            "metrics": {
                str(name): compact_metric(
                    value,
                    ((previous or {}).get("metrics", {}) | previous_other_metrics).get(name, {}),
                    as_of_date,
                )
                for name, value in {**row.get("metrics", {}), **other_metrics}.items()
            },
        }

    personal_date_range, personal_rows = extract_rows(source_path)
    previous_source_path = source_path.parent / previous_date.strftime("%m%d.xlsx")
    previous_personal_rows: list[dict[str, Any]] = []
    if previous_source_path.exists():
        _, previous_personal_rows = extract_rows(previous_source_path)
        apply_previous_rates(personal_rows, previous_personal_rows)
    else:
        for row in personal_rows:
            row["previous_overall_rate"] = None
            row["overall_rate_dod"] = None

    previous_person_by_id = {
        row["employee_id"]: row for row in previous_personal_rows if row.get("employee_id")
    }
    previous_person_by_key = {
        (row["store"], row["role"], row["name"]): row for row in previous_personal_rows
    }
    person_awards_by_id: dict[str, dict[str, Any]] = {}
    if phone_summary_path and phone_summary_path.exists():
        phone_summary = json.loads(phone_summary_path.read_text(encoding="utf-8"))
        person_awards_by_id = {
            str(item.get("emp_id")): item
            for item in phone_summary.get("person_awards", [])
            if item.get("emp_id")
        }
    personal = []
    for row in personal_rows:
        metrics = {item["label"]: compact_metric(item, as_of_date=as_of_date) for item in row.get("metrics", [])}
        person_award = person_awards_by_id.get(str(row.get("employee_id", "")), {})
        personal.append({
            "store": normalize_store(str(row.get("store", ""))),
            "role": row.get("role"),
            "category": row.get("category"),
            "name": mask_name(row.get("name")),
            "rank": numeric(row.get("rank")),
            "overall_rate": numeric(row.get("overall_rate")),
            "overall_rate_dod": numeric(row.get("overall_rate_dod")),
            "rank_dod": None,
            "phone_award_actual": numeric(person_award.get("actual_total")),
            "phone_award_projected": numeric(person_award.get("projected")),
            "phone_award_rank": numeric(person_award.get("rank")),
            "phone_award_eligible": person_award.get("award"),
            "insurance_attach_rate": numeric(row.get("insurance_attach_rate")),
            "metrics": metrics,
        })

        previous_person = previous_person_by_id.get(row.get("employee_id")) or previous_person_by_key.get((row.get("store"), row.get("role"), row.get("name")))
        if previous_person and numeric(row.get("rank")) is not None and numeric(previous_person.get("rank")) is not None:
            personal[-1]["rank_dod"] = int(numeric(previous_person["rank"]) - numeric(row["rank"]))

    return {
        "version": 1,
        "visibility": "private-local-preview",
        "report_date": as_of_date.isoformat(),
        "report_run_date": report_run_date.isoformat(),
        "data_as_of_date": as_of_date.isoformat(),
        "source_as_of_date": as_of_date.isoformat(),
        "source_file": source_path.name,
        "previous_report_date": previous_date.isoformat() if previous_report else None,
        "source_date_range": report.get("source_date_range") or personal_date_range,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "aggregate": compact_store(report.get("aggregate", {})),
        "stores": [compact_store(row) for row in report.get("records", [])],
        "personal": personal,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, default=Path("/Users/liamlu/Library/CloudStorage/OneDrive-個人/TWM每日戰報"))
    parser.add_argument("--outputs-dir", type=Path, default=Path("/Users/liamlu/Downloads/liam-agent/report-automation/outputs"))
    parser.add_argument("--site-data-dir", type=Path, default=Path("/Users/liamlu/Downloads/liam-agent/github-pages-liamlu/data"))
    parser.add_argument("--today-report-path", type=Path, default=Path("/Users/liamlu/Downloads/liam-agent/report-automation/work/today_report_data.json"))
    parser.add_argument("--private-kpi-output", type=Path, default=Path("/Users/liamlu/Downloads/liam-agent/github-pages-liamlu/private-data/kpi-battle-latest.json"))
    parser.add_argument("--report-run-date", help="Mail/report execution date (YYYY-MM-DD)")
    parser.add_argument("--data-cutoff-date", help="Formal KPI/awards data date (YYYY-MM-DD)")
    args = parser.parse_args()

    if not args.today_report_path.exists():
        raise RuntimeError(f"today_report_data.json 不存在：{args.today_report_path}")
    today_report = json.loads(args.today_report_path.read_text(encoding="utf-8"))
    report_run_date, data_cutoff_date = resolve_date_contract(
        today_report,
        args.report_run_date,
        args.data_cutoff_date,
    )
    report_run_date_iso = report_run_date.isoformat()
    data_cutoff_date_iso = data_cutoff_date.isoformat()
    source_file = canonical_input_source_file(today_report)

    args.site_data_dir.mkdir(parents=True, exist_ok=True)
    momentum = extract_momentum(args.source_dir)
    (args.site_data_dir / "daily-momentum.json").write_text(
        json.dumps(momentum, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    summary_path = args.outputs_dir / "phone_awards_update_summary.json"
    if summary_path.exists():
        phone = extract_phone_awards(
            summary_path,
            report_run_date_iso,
            data_cutoff_date_iso,
        )
        (args.site_data_dir / "phone-awards-latest.json").write_text(
            json.dumps(phone, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    phone_battle_written = False
    if summary_path.exists():
        args.private_kpi_output.parent.mkdir(parents=True, exist_ok=True)
        phone_battle = extract_phone_awards_battle(
            summary_path,
            report_run_date_iso,
            data_cutoff_date_iso,
        )
        (args.private_kpi_output.parent / "phone-awards-battle-latest.json").write_text(
            json.dumps(phone_battle, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        phone_battle_written = True

    kpi_battle_written = False
    if args.today_report_path.exists():
        args.private_kpi_output.parent.mkdir(parents=True, exist_ok=True)
        supplemental_path = args.outputs_dir / f"supplemental_daily_report_{report_run_date_iso}.json"
        kpi_battle = extract_kpi_battle(
            args.today_report_path,
            summary_path if summary_path.exists() else None,
            supplemental_path if supplemental_path.exists() else None,
            report_run_date_iso,
            data_cutoff_date_iso,
        )
        args.private_kpi_output.write_text(
            json.dumps(kpi_battle, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        kpi_battle_written = True

    print(json.dumps({
        "momentum_stores": len(momentum["stores"]),
        "momentum_dates": len(momentum["latest_source_by_date"]),
        "source_files": len(momentum["source_files"]),
        "report_run_date": report_run_date_iso,
        "data_cutoff_date": data_cutoff_date_iso,
        "site_data_dir": str(args.site_data_dir),
        "private_kpi_output": str(args.private_kpi_output) if kpi_battle_written else None,
        "private_phone_awards_output": str(args.private_kpi_output.parent / "phone-awards-battle-latest.json") if phone_battle_written else None,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
